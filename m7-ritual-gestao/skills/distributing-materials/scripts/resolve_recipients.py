"""
resolve_recipients.py — Resolve destinatarios DM via Calendario-de-Rituais.xlsx.

Le a sheet 'Calendario' e busca a linha {Vertical, Nivel} correspondente.
Retorna JSON com Gestor + Participantes + Lider Direto (User IDs U... do Slack).

Schema esperado do XLSX (apos extensao manual conforme calendar-schema.md):
    Coluna existente    | Tipo
    --------------------+------
    Vertical            | str (Investimentos, Credito, Universo, Consorcios, Seguros)
    Sigla               | str (INV, CRE, UNI, CON, SEG)
    Nivel               | str ("N3 - Operacional", "N2 - Tatico", ou apenas "N3"/"N2")
    Frequencia          | str (Semanal, Mensal)
    ...                 | (outras colunas existentes preservadas)
    --------------------+------
    Coluna NOVA         | Tipo
    --------------------+------
    Gestor-User-ID      | str (U... do Slack)
    Participantes-Nomes | str "; "-separated
    Participantes-User-IDs | str "; "-separated (U...)
    Lider-Direto-User-ID | str (U... ou vazio)

Uso CLI:
    python resolve_recipients.py \\
      --calendar-path /path/to/Calendario-de-Rituais.xlsx \\
      --vertical consorcios \\
      --nivel N3 \\
      [--subnivel wl] \\
      [--include-escalacao]

Output: JSON em stdout.

Exit codes:
    0 = OK, linha encontrada e colunas preenchidas
    2 = erro de invocacao (path invalido, openpyxl ausente, etc)
    3 = linha nao encontrada
    4 = colunas novas ausentes ou vazias
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print(
        json.dumps({"ok": False, "error": "openpyxl nao instalado. Run: pip install openpyxl"}),
        file=sys.stdout,
    )
    sys.exit(2)


# Mapping vertical lowercase -> nome exato como aparece no XLSX (case-sensitive na coluna)
VERTICAL_MAP = {
    "investimentos": "Investimentos",
    "credito": "Crédito",
    "universo": "Universo",
    "consorcios": "Consórcios",
    "seguros": "Seguros",
    "pj2": "PJ2",
}

# Colunas NOVAS que precisam estar presentes (case-sensitive no header)
REQUIRED_NEW_COLUMNS = (
    "Gestor-User-ID",
    "Participantes-Nomes",
    "Participantes-User-IDs",
)
# Coluna opcional (usada so quando --include-escalacao em post_ritual com escalacao)
ESCALACAO_COLUMN = "Lider-Direto-User-ID"
# Coluna opcional pos-ritual (obrigatoria em post_ritual; pode estar vazia em verticais sem canal)
CANAL_COLUMN = "Canal-Vertical-ID"


def _normalize_nivel(n: str) -> str:
    """Aceita 'N3', 'N3 - Operacional', 'n3' etc. Retorna 'N1'..'N5'."""
    s = (n or "").strip().upper()
    # extrair prefixo Nx
    if s.startswith("N") and len(s) >= 2 and s[1].isdigit():
        return s[:2]
    return s


def _norm_vertical_for_match(v: str) -> str:
    """Normaliza vertical pra match: lowercase + remove acento."""
    import unicodedata
    s = (v or "").strip().lower()
    nfkd = unicodedata.normalize("NFKD", s)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def _norm_header(h) -> str:
    """Normaliza nome de coluna pra match case/acento-insensitive (2026-06-18).

    O header do XLSX era resolvido case/acento-sensitive -> 'VERTICAL', 'vertical'
    ou 'Nível' (acento natural PT) quebravam com 'Coluna ausente' criptico. Agora
    a busca de coluna passa por aqui nos dois lados.
    """
    import unicodedata
    s = str(h or "").strip().lower()
    nfkd = unicodedata.normalize("NFKD", s)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def _split_semicolon(raw: str) -> list[str]:
    if not raw:
        return []
    return [x.strip() for x in str(raw).split(";") if x.strip()]


def resolve(
    calendar_path: Path,
    vertical: str,
    nivel: str,
    *,
    subnivel: str = "",
    include_escalacao: bool = False,
) -> dict:
    """Resolve destinatarios a partir do XLSX.

    Returns:
        {
          "ok": bool,
          "vertical": "consorcios",
          "nivel": "N3",
          "subnivel": "" or "wl",
          "gestor": {"name": str, "user_id": "U..."},
          "participantes": [{"name", "user_id"}],
          "lider_direto": {"name", "user_id"} | None,
          "row_meta": {"sheet_row": int, "...": "..."},
          "error": str (se ok=false)
        }
    """
    if not calendar_path.exists():
        return {"ok": False, "error": f"Calendario nao encontrado: {calendar_path}"}

    try:
        wb = load_workbook(filename=str(calendar_path), data_only=True, read_only=True)
    except Exception as exc:
        return {"ok": False, "error": f"Falha ao abrir XLSX: {exc}"}

    sheet_name = "Calendario"
    if sheet_name not in wb.sheetnames:
        # tentar variantes
        for cand in ("Calendario", "Calendário", "calendario"):
            if cand in wb.sheetnames:
                sheet_name = cand
                break
        else:
            return {"ok": False, "error": f"Sheet 'Calendario' nao encontrada. Sheets: {wb.sheetnames}"}

    ws = wb[sheet_name]

    # Header na linha 1
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return {"ok": False, "error": "XLSX sem header"}

    # Index normalizado (case/acento-insensitive). Acesso sempre via _h(nome).
    header_idx = {_norm_header(h): i for i, h in enumerate(header) if h is not None}

    def _h(name) -> int | None:
        return header_idx.get(_norm_header(name))

    # Validar colunas novas
    missing_new = [c for c in REQUIRED_NEW_COLUMNS if _h(c) is None]
    if missing_new:
        return {
            "ok": False,
            "error": (
                f"Colunas novas ausentes no XLSX: {missing_new}. "
                f"Estenda Calendario-de-Rituais.xlsx conforme calendar-schema.md "
                f"(adicione: {', '.join(REQUIRED_NEW_COLUMNS)} + opcional Lider-Direto-User-ID)."
            ),
            "missing_columns": missing_new,
        }

    # Validar coluna Vertical e Nivel
    if _h("Vertical") is None:
        return {"ok": False, "error": "Coluna 'Vertical' ausente no XLSX"}
    if _h("Nivel") is None:
        return {"ok": False, "error": "Coluna 'Nivel' ausente no XLSX"}

    vert_target = _norm_vertical_for_match(vertical)
    niv_target = _normalize_nivel(nivel)
    sub_target = (subnivel or "").strip().lower()
    vert_col, niv_col = _h("Vertical"), _h("Nivel")
    sub_col = _h("Subnivel")

    # Coletar TODAS as linhas que matcheam Vertical+Nivel
    candidates = []  # lista de (row_index, row_tuple, subnivel_str)
    for i, row in enumerate(rows, start=2):
        v_cell = row[vert_col] if row[vert_col] else ""
        n_cell = row[niv_col] if row[niv_col] else ""
        if _norm_vertical_for_match(str(v_cell)) != vert_target:
            continue
        if _normalize_nivel(str(n_cell)) != niv_target:
            continue
        sub_cell = ""
        if sub_col is not None:
            raw = row[sub_col]
            sub_cell = str(raw).strip().lower() if raw else ""
        candidates.append((i, row, sub_cell))

    if not candidates:
        return {
            "ok": False,
            "error": f"Linha nao encontrada: vertical={vertical} nivel={niv_target}",
            "code": "row_not_found",
        }

    # Filtrar por Subnivel
    if sub_target:
        filtered = [c for c in candidates if c[2] == sub_target]
        if not filtered:
            disponiveis = sorted({c[2] for c in candidates if c[2]}) or ["(nenhum)"]
            return {
                "ok": False,
                "error": (
                    f"Subnivel '{sub_target}' nao encontrado em {vertical} {niv_target}. "
                    f"Subniveis disponiveis: {disponiveis}"
                ),
                "code": "subnivel_not_found",
            }
        if len(filtered) > 1:
            return {
                "ok": False,
                "error": f"Multiplas linhas para {vertical} {niv_target} subnivel={sub_target}: {[c[0] for c in filtered]}",
                "code": "ambiguous_subnivel_match",
            }
        match_idx, match_row, _ = filtered[0]
    else:
        # Sem subnivel: aceita unica linha sem subnivel preenchido, OU primeira sem subnivel
        consolidados = [c for c in candidates if not c[2]]
        com_subnivel = [c for c in candidates if c[2]]
        if consolidados and not com_subnivel:
            if len(consolidados) > 1:
                return {
                    "ok": False,
                    "error": f"Multiplas linhas consolidadas para {vertical} {niv_target}: {[c[0] for c in consolidados]}",
                    "code": "ambiguous_consolidated",
                }
            match_idx, match_row, _ = consolidados[0]
        elif com_subnivel and not consolidados:
            disponiveis = sorted({c[2] for c in com_subnivel})
            return {
                "ok": False,
                "error": (
                    f"Vertical {vertical} {niv_target} esta dividida por subnivel ({disponiveis}). "
                    f"Informe --subnivel."
                ),
                "code": "subnivel_required",
                "subniveis_disponiveis": disponiveis,
            }
        else:
            # Tem ambos (consolidado + split) — caso hibrido raro; preferir consolidado
            if len(consolidados) > 1:
                return {
                    "ok": False,
                    "error": f"Multiplas linhas consolidadas para {vertical} {niv_target}",
                    "code": "ambiguous_consolidated",
                }
            match_idx, match_row, _ = consolidados[0]

    # Extrair colunas
    gestor_id = match_row[_h("Gestor-User-ID")]
    part_names_raw = match_row[_h("Participantes-Nomes")]
    part_ids_raw = match_row[_h("Participantes-User-IDs")]
    lider_id = match_row[_h(ESCALACAO_COLUMN)] if _h(ESCALACAO_COLUMN) is not None else None
    canal_id = match_row[_h(CANAL_COLUMN)] if _h(CANAL_COLUMN) is not None else None
    condutor_name = match_row[_h("Condutor")] if _h("Condutor") is not None else None

    # Validar conteudo
    empty_cells = []
    if not gestor_id:
        empty_cells.append("Gestor-User-ID")
    if not part_ids_raw:
        empty_cells.append("Participantes-User-IDs")
    if empty_cells:
        return {
            "ok": False,
            "error": (
                f"Linha encontrada (vertical={vertical} nivel={niv_target}) mas com celulas vazias: "
                f"{empty_cells}. Preencha-as no Calendario-de-Rituais.xlsx."
            ),
            "code": "empty_cells",
            "empty_cells": empty_cells,
            "sheet_row": match_idx,
        }

    part_names = _split_semicolon(part_names_raw)
    part_ids = _split_semicolon(part_ids_raw)
    if len(part_names) != len(part_ids):
        return {
            "ok": False,
            "error": (
                f"Mismatch entre 'Participantes-Nomes' ({len(part_names)}) e "
                f"'Participantes-User-IDs' ({len(part_ids)}) na linha {match_idx}. "
                f"Devem ter o mesmo numero de itens separados por ';'."
            ),
            "code": "mismatch_participantes",
            "sheet_row": match_idx,
        }

    participantes = [{"name": n, "user_id": uid} for n, uid in zip(part_names, part_ids)]

    out = {
        "ok": True,
        "vertical": vertical,
        "nivel": niv_target,
        "subnivel": subnivel or "",
        "gestor": {"name": condutor_name or "(condutor nao informado no XLSX)", "user_id": str(gestor_id).strip()},
        "participantes": participantes,
        "lider_direto": None,
        "canal_id": str(canal_id).strip() if canal_id else None,
        "row_meta": {"sheet_row": match_idx, "sheet_name": sheet_name},
    }

    if include_escalacao and lider_id:
        out["lider_direto"] = {"name": "(lider direto)", "user_id": str(lider_id).strip()}

    return out


def main():
    p = argparse.ArgumentParser(description="Resolve recipients from Calendario-de-Rituais.xlsx")
    p.add_argument("--calendar-path", required=True)
    p.add_argument("--vertical", required=True)
    p.add_argument("--nivel", required=True, help="N1/N2/N3")
    p.add_argument("--subnivel", default="")
    p.add_argument("--include-escalacao", action="store_true")
    args = p.parse_args()

    res = resolve(
        Path(args.calendar_path),
        args.vertical,
        args.nivel,
        subnivel=args.subnivel,
        include_escalacao=args.include_escalacao,
    )

    print(json.dumps(res, ensure_ascii=False, indent=2))
    if not res.get("ok"):
        code = res.get("code", "")
        if code == "row_not_found":
            sys.exit(3)
        if code in ("empty_cells", "mismatch_participantes"):
            sys.exit(4)
        sys.exit(2)
    sys.exit(0)


if __name__ == "__main__":
    main()
