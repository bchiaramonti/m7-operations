"""
_lib.py — Utilitarios compartilhados pelos scripts de managing-action-plan.

Este modulo NAO e invocado diretamente. Os scripts da skill o importam via:

    import sys, os
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from _lib import normalize_date, canonical_row, hash_row, ...

Conteudo:
    - Constantes de schema (colunas, enums, hierarquia)
    - Parsing de datas BR ("02/abr" -> datetime) + normalizacao
    - Canonicalizacao de linha + hash SHA-256
    - IO xlsx (load, find_data_range, write_clickup_id_column)
    - Resolucao de hierarquia via No. (1.1.1 -> parent 1.1)
"""

from __future__ import annotations

import datetime as _dt
import hashlib
import re
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError as e:
    raise SystemExit(
        "ERRO: openpyxl nao instalado. Execute: pip3 install openpyxl"
    ) from e


# ---------------------------------------------------------------------------
# Schema constants
# ---------------------------------------------------------------------------

# Colunas esperadas (em ordem). Header esta na row HEADER_ROW (4).
# Coluna A e margem visual (vazia).
HEADER_ROW = 4
FIRST_DATA_ROW = 5

# Mapping nome canonico -> labels possiveis no header (para tolerar variacoes)
COLUMNS = {
    "no":          ["No.", "No", "Numero", "Número"],
    "tipo":        ["Tipo"],
    "etapa":       ["Etapa", "Atividade", "Item"],
    "responsavel": ["Responsável", "Responsavel", "Owner"],
    "inicio_plan": ["Início Planejado", "Inicio Planejado", "Início Plan.", "Inicio Plan."],
    "fim_plan":    ["Fim Planejado", "Fim Plan."],
    "inicio_real": ["Início Real", "Inicio Real"],
    "fim_real":    ["Fim Real"],
    "status":      ["Status"],
    "entregavel":  ["Entregável", "Entregavel", "Deliverable"],
    "clickup_id":  ["ClickUp ID", "Clickup ID", "CU ID"],  # adicionada por esta skill
}

# Ordem canonica das colunas para escrita (e a ordem visual esperada)
CANONICAL_COLUMN_ORDER = [
    "no", "tipo", "etapa", "responsavel",
    "inicio_plan", "fim_plan", "inicio_real", "fim_real",
    "status", "entregavel", "clickup_id",
]

# Tipos de linha (coluna C)
TIPOS_VALIDOS = {"Fase", "Ação", "Acao", "Etapas da Ação", "Etapas da Acao"}
TIPO_FASE = "Fase"
TIPO_ACAO = "Ação"
TIPO_ETAPA = "Etapas da Ação"

# Status enum local + map default para custom statuses do ClickUp
STATUS_LOCAL_VALIDOS = {"not_started", "in_progress", "blocked", "done", ""}
STATUS_MAP_DEFAULT = {
    "not_started": "to do",
    "in_progress": "in progress",
    "blocked": "blocked",
    "done": "complete",
}

# Meses em portugues (abreviacao usada no formato "02/abr")
MESES_BR = {
    "jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
    "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12,
}

NO_PATTERN = re.compile(r"^\d+(\.\d+)*$")
DATE_BR_PATTERN = re.compile(r"^(\d{1,2})/([a-z]{3})(?:/(\d{4}))?$", re.IGNORECASE)
DATE_NUMERIC_PATTERN = re.compile(r"^(\d{1,2})/(\d{1,2})(?:/(\d{4}))?$")


# ---------------------------------------------------------------------------
# Date normalization
# ---------------------------------------------------------------------------

def normalize_date(value: Any, default_year: int | None = None) -> _dt.datetime | None:
    """
    Normaliza um valor de celula para datetime ou None.

    Aceita:
        - datetime / date -> datetime (00:00)
        - "02/abr" ou "02/abr/2026" -> datetime
        - "02/04" ou "02/04/2026" -> datetime (assume formato dd/mm)
        - "" / None / "—" -> None
        - ISO "2026-04-02" -> datetime

    `default_year` e usado quando a string nao traz ano. Se None, usa
    o ano corrente (datetime.now().year).
    """
    if value is None:
        return None
    if isinstance(value, _dt.datetime):
        return value
    if isinstance(value, _dt.date):
        return _dt.datetime(value.year, value.month, value.day)
    if not isinstance(value, str):
        return None
    s = value.strip()
    if s in ("", "—", "-", "--"):
        return None

    # ISO YYYY-MM-DD
    try:
        return _dt.datetime.strptime(s, "%Y-%m-%d")
    except ValueError:
        pass

    year = default_year or _dt.datetime.now().year

    # "02/abr" ou "02/abr/2026"
    m = DATE_BR_PATTERN.match(s)
    if m:
        day = int(m.group(1))
        month_abbr = m.group(2).lower()
        if month_abbr not in MESES_BR:
            return None
        month = MESES_BR[month_abbr]
        y = int(m.group(3)) if m.group(3) else year
        try:
            return _dt.datetime(y, month, day)
        except ValueError:
            return None

    # "02/04" ou "02/04/2026" (assume dd/mm — convencao BR)
    m = DATE_NUMERIC_PATTERN.match(s)
    if m:
        day = int(m.group(1))
        month = int(m.group(2))
        y = int(m.group(3)) if m.group(3) else year
        try:
            return _dt.datetime(y, month, day)
        except ValueError:
            return None

    return None


def date_to_iso(value: Any, default_year: int | None = None) -> str:
    """Converte valor de data para ISO YYYY-MM-DD ou string vazia se None/invalido."""
    dt = normalize_date(value, default_year)
    return dt.strftime("%Y-%m-%d") if dt else ""


# ---------------------------------------------------------------------------
# Hierarchy resolution
# ---------------------------------------------------------------------------

def parent_no(no: str) -> str:
    """
    Retorna o `No.` do pai dado um `No.` filho.

    Exemplos:
        parent_no("1.1.1") -> "1.1"
        parent_no("1.1")   -> "1"
        parent_no("1")     -> ""  (root)
        parent_no("")      -> ""

    Nao valida se o pai existe na tabela (e responsabilidade do chamador).
    """
    if not no or "." not in no:
        return ""
    return no.rsplit(".", 1)[0]


def hierarchy_level(no: str) -> int:
    """Retorna o nivel hierarquico baseado no No. (1=Fase, 2=Acao, 3=Etapa)."""
    if not no:
        return 0
    return no.count(".") + 1


# ---------------------------------------------------------------------------
# Row canonicalization + hashing
# ---------------------------------------------------------------------------

# Campos incluidos no hash (excluem clickup_id que e mapping, nao conteudo)
HASH_FIELDS = [
    "no", "tipo", "etapa", "responsavel",
    "inicio_plan", "fim_plan", "inicio_real", "fim_real",
    "status", "entregavel",
]


def canonical_row(row: dict[str, Any], default_year: int | None = None) -> dict[str, str]:
    """
    Canonicaliza uma linha para hashing/comparacao deterministica.

    - Strings: trim, normaliza unicode whitespace
    - Datas: ISO YYYY-MM-DD ou ""
    - None: ""
    - clickup_id: preservado mas nao incluido em HASH_FIELDS
    """
    out: dict[str, str] = {}
    for field in CANONICAL_COLUMN_ORDER:
        v = row.get(field, "")
        if v is None:
            out[field] = ""
            continue
        if field in ("inicio_plan", "fim_plan", "inicio_real", "fim_real"):
            out[field] = date_to_iso(v, default_year)
        else:
            s = str(v).strip()
            # Normalizar marcadores de "vazio"
            if s in ("—", "-", "--"):
                s = ""
            out[field] = s
    return out


def hash_row(row: dict[str, Any], default_year: int | None = None) -> str:
    """SHA-256 hex da concatenacao canonica dos HASH_FIELDS."""
    canon = canonical_row(row, default_year)
    payload = "|".join(canon[f] for f in HASH_FIELDS)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def hash_table(rows: list[dict[str, Any]], default_year: int | None = None) -> str:
    """Hash agregado da tabela inteira (ordenado por No. para estabilidade)."""
    sorted_rows = sorted(rows, key=lambda r: _no_sort_key(r.get("no", "")))
    h = hashlib.sha256()
    for r in sorted_rows:
        h.update(hash_row(r, default_year).encode("utf-8"))
    return h.hexdigest()


def _no_sort_key(no: str) -> tuple:
    """Sort key para `No.` hierarquico: '1.10.2' > '1.2.5' (numerico, nao lexico)."""
    if not no:
        return ()
    try:
        return tuple(int(p) for p in no.split("."))
    except ValueError:
        return (0,)  # fallback para Nos. malformados


# ---------------------------------------------------------------------------
# XLSX IO
# ---------------------------------------------------------------------------

class CronogramaXLSX:
    """
    Wrapper para o cronograma xlsx. Encapsula:
      - load + sheet detection
      - mapeamento header -> coluna canonica
      - leitura de linhas como dicts canonicos
      - escrita preservando formatacao
      - garantia da coluna ClickUp ID

    Uso:
        cron = CronogramaXLSX(path)
        cron.load()
        rows = cron.read_rows()
        cron.ensure_clickup_id_column()
        cron.write_clickup_id(row_index=5, clickup_id="86abc")
        cron.save()
    """

    def __init__(self, path: str | Path):
        self.path = Path(path)
        self.wb: openpyxl.Workbook | None = None
        self.ws = None
        self.header_row = HEADER_ROW
        self.first_data_row = FIRST_DATA_ROW
        # Map: nome canonico -> indice de coluna (1-based, openpyxl)
        self.col_index: dict[str, int] = {}

    def load(self) -> None:
        if not self.path.exists():
            raise FileNotFoundError(f"Cronograma nao encontrado: {self.path}")
        self.wb = openpyxl.load_workbook(self.path, data_only=False)
        # Pega a primeira sheet por convencao (esperado: "Cronograma Detalhado")
        self.ws = self.wb.active
        self._detect_columns()

    def _detect_columns(self) -> None:
        """Mapeia cada coluna canonica para o indice real no header."""
        if self.ws is None:
            raise RuntimeError("ws nao carregado — chame load() primeiro")
        # Procura header_row procurando por celula que case com algum label conhecido
        for col in range(1, self.ws.max_column + 1):
            value = self.ws.cell(self.header_row, col).value
            if not value or not isinstance(value, str):
                continue
            label = value.strip()
            for canon, labels in COLUMNS.items():
                if label in labels:
                    self.col_index[canon] = col
                    break
        # Validar minimo: precisa ter no, tipo, etapa
        for required in ("no", "tipo", "etapa"):
            if required not in self.col_index:
                raise ValueError(
                    f"Coluna obrigatoria '{required}' nao encontrada no header "
                    f"da row {self.header_row}. Header detectado: "
                    f"{[self.ws.cell(self.header_row, c).value for c in range(1, self.ws.max_column + 1)]}"
                )

    def last_data_row(self) -> int:
        """Detecta a ultima linha com dados (coluna `no` preenchida)."""
        if self.ws is None or "no" not in self.col_index:
            return self.first_data_row - 1
        col = self.col_index["no"]
        last = self.first_data_row - 1
        for r in range(self.first_data_row, self.ws.max_row + 1):
            v = self.ws.cell(r, col).value
            if v is not None and str(v).strip():
                last = r
        return last

    def read_rows(self) -> list[dict[str, Any]]:
        """
        Le todas as linhas de dados como lista de dicts canonicos.
        Cada dict inclui:
            - Os campos canonicos (no, tipo, etapa, ...)
            - "_row_index": numero da linha no xlsx (para writeback)
        """
        if self.ws is None:
            raise RuntimeError("ws nao carregado")
        rows: list[dict[str, Any]] = []
        for r in range(self.first_data_row, self.last_data_row() + 1):
            row: dict[str, Any] = {"_row_index": r}
            for canon, col in self.col_index.items():
                row[canon] = self.ws.cell(r, col).value
            # So inclui se tiver `no` preenchido
            if row.get("no"):
                rows.append(row)
        return rows

    def ensure_clickup_id_column(self) -> int:
        """
        Garante que a coluna `ClickUp ID` existe. Se nao existir, adiciona
        depois da ULTIMA coluna preenchida no header (preserva margem A
        e qualquer outra coluna intencionalmente vazia entre as preenchidas).
        Retorna o indice da coluna.
        """
        if self.ws is None:
            raise RuntimeError("ws nao carregado")
        if "clickup_id" in self.col_index:
            return self.col_index["clickup_id"]
        # Encontrar a maior coluna com header preenchido
        last_filled = 0
        for col in range(1, self.ws.max_column + 1):
            v = self.ws.cell(self.header_row, col).value
            if v is not None and isinstance(v, str) and v.strip():
                last_filled = col
        if last_filled == 0:
            # Header completamente vazio — improvavel, mas defensivo
            new_col = 2  # nunca usa coluna A (margem)
        else:
            new_col = last_filled + 1
        cell = self.ws.cell(self.header_row, new_col, "ClickUp ID")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        self.col_index["clickup_id"] = new_col
        return new_col

    def write_clickup_id(self, row_index: int, clickup_id: str) -> None:
        if self.ws is None or "clickup_id" not in self.col_index:
            raise RuntimeError("ws ou coluna clickup_id ausente")
        self.ws.cell(row_index, self.col_index["clickup_id"], clickup_id)

    def write_cell(self, row_index: int, field: str, value: Any) -> None:
        """Escreve um valor em campo canonico (datas viram datetime)."""
        if self.ws is None or field not in self.col_index:
            raise RuntimeError(f"ws ou coluna '{field}' ausente")
        if field in ("inicio_plan", "fim_plan", "inicio_real", "fim_real") and value:
            dt = normalize_date(value)
            if dt:
                self.ws.cell(row_index, self.col_index[field], dt)
                return
        self.ws.cell(row_index, self.col_index[field], value)

    def append_row(self, row: dict[str, Any]) -> int:
        """Append uma nova linha apos a ultima. Retorna o row_index."""
        if self.ws is None:
            raise RuntimeError("ws nao carregado")
        new_r = self.last_data_row() + 1
        for field, col in self.col_index.items():
            v = row.get(field)
            if v is None:
                continue
            if field in ("inicio_plan", "fim_plan", "inicio_real", "fim_real"):
                dt = normalize_date(v)
                if dt:
                    self.ws.cell(new_r, col, dt)
                    continue
            self.ws.cell(new_r, col, v)
        return new_r

    def delete_row(self, row_index: int) -> None:
        """Remove uma linha (shifts up)."""
        if self.ws is None:
            raise RuntimeError("ws nao carregado")
        self.ws.delete_rows(row_index, 1)

    def save(self, dest: str | Path | None = None) -> Path:
        if self.wb is None:
            raise RuntimeError("wb nao carregado")
        target = Path(dest) if dest else self.path
        target.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(target)
        return target


# ---------------------------------------------------------------------------
# Sync state sidecar IO
# ---------------------------------------------------------------------------

import json as _json  # noqa: E402


SYNC_STATE_FILENAME = ".sync-state.json"

SYNC_STATE_DEFAULT = {
    "schema_version": 1,
    "last_sync": "",            # ISO 8601
    "last_sync_hash": "",       # hash da tabela apos ultimo sync ok
    "sync_pending": False,      # true se algum push falhou parcialmente
    "pending_ops": [],          # lista de ops pendentes (formato livre, Claude usa)
    "history": [],              # ultimas N entries de sync (max 20)
}

_HISTORY_MAX = 20


def sync_state_path(status_report_dir: str | Path) -> Path:
    return Path(status_report_dir) / SYNC_STATE_FILENAME


def read_sync_state(status_report_dir: str | Path) -> dict:
    """Le sync state; retorna SYNC_STATE_DEFAULT se nao existe."""
    p = sync_state_path(status_report_dir)
    if not p.exists():
        return dict(SYNC_STATE_DEFAULT)
    try:
        data = _json.loads(p.read_text(encoding="utf-8"))
    except _json.JSONDecodeError:
        # Sidecar corrompido — tratar como inexistente, mas nao apagar
        return dict(SYNC_STATE_DEFAULT)
    # Merge com defaults para tolerar campos novos
    merged = dict(SYNC_STATE_DEFAULT)
    merged.update(data)
    return merged


def write_sync_state(status_report_dir: str | Path, state: dict) -> Path:
    p = sync_state_path(status_report_dir)
    p.parent.mkdir(parents=True, exist_ok=True)
    # Trim history
    if "history" in state and isinstance(state["history"], list):
        state["history"] = state["history"][-_HISTORY_MAX:]
    p.write_text(_json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")
    return p


def update_sync_state(status_report_dir: str | Path, **fields) -> dict:
    """Le, atualiza campos e regrava. Retorna o estado atualizado."""
    state = read_sync_state(status_report_dir)
    state.update(fields)
    write_sync_state(status_report_dir, state)
    return state


def append_sync_history(status_report_dir: str | Path, entry: dict) -> dict:
    """Anexa um entry ao history (trim automatico)."""
    state = read_sync_state(status_report_dir)
    state.setdefault("history", []).append(entry)
    write_sync_state(status_report_dir, state)
    return state
