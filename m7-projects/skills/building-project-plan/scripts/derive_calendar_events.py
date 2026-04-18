#!/usr/bin/env python3
"""
derive_calendar_events.py — Deriva events[] do calendario a partir do
Cronograma.xlsx (linhas Tipo=Fase + marcos) + rituais coletados na Fase B.

Tipos de evento gerados:
    fase           — Inicio/fim de cada Tipo=Fase (uma linha "Inicio Fase X" no inicio_plan,
                     uma linha "Fim Fase X" no fim_plan, mas opcional)
    marco          — Marcos definidos em data["roadmap"]["milestones"]
    discovery      — Sessoes de discovery (data["comunicacao"]["rituais"] com type="discovery")
    validacao      — Validacoes
    checkin        — Check-ins recorrentes
    status-report  — Status reports
    handoff        — Handoffs

Para rituais recorrentes (data["recurring_rituals"]), gera ocorrencias
em datas calculadas conforme freq + start + end.

Uso:
    python3 derive_calendar_events.py --input plan.json --xlsx Cronograma.xlsx --output events.json
    python3 derive_calendar_events.py --input ... --inline-into plan.json   # injeta direto em plan.events

Saida (--output):
    JSON array de events conforme schema esperado por calendario.tmpl.html:
    [{"d":"YYYY-MM-DD","type":"...","label":"...","who":"...","dur":"...","out":"..."}, ...]
"""

import argparse
import datetime as dt
import json
import os
import sys
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _lib import parse_date  # noqa: E402

try:
    import openpyxl
except ImportError:
    print("ERRO: openpyxl nao instalado. Execute: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)


HEADER_ROW = 4
FIRST_DATA_ROW = 5

# Map de freq textual -> dias entre ocorrencias (apenas heuristica para rituais recorrentes)
FREQ_DAYS = {
    "diario": 1,
    "semanal": 7,
    "quinzenal": 14,
    "mensal": 30,
}


def fases_from_xlsx(xlsx_path: Path) -> list[dict]:
    """Le linhas Tipo=Fase do xlsx e gera events de inicio/fim."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    ws = wb.active
    events = []
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        no = ws.cell(r, 2).value
        tipo = ws.cell(r, 3).value
        etapa = ws.cell(r, 4).value
        responsavel = ws.cell(r, 5).value or ""
        ini = ws.cell(r, 6).value
        fim = ws.cell(r, 7).value
        if not no or str(tipo).strip() != "Fase":
            continue
        # Apenas Fases raiz (sem ponto) viram events de calendario; subfases sao detalhes
        if "." in str(no):
            continue
        ini_d = parse_date(ini)
        fim_d = parse_date(fim)
        if ini_d:
            events.append({
                "d": ini_d.strftime("%Y-%m-%d"),
                "type": "fase",
                "label": f"Inicio: {etapa}",
                "who": str(responsavel),
                "dur": "",
                "out": "",
            })
        if fim_d:
            events.append({
                "d": fim_d.strftime("%Y-%m-%d"),
                "type": "fase",
                "label": f"Fim: {etapa}",
                "who": str(responsavel),
                "dur": "",
                "out": "",
            })
    return events


def milestones_to_events(milestones: list[dict]) -> list[dict]:
    """Converte milestones do roadmap em events tipo 'marco'."""
    events = []
    for m in milestones:
        d = parse_date(m.get("date_iso") or m.get("date"))
        if not d:
            continue
        events.append({
            "d": d.strftime("%Y-%m-%d"),
            "type": "marco",
            "label": f'Marco: {m.get("h4", "")}',
            "who": "",
            "dur": "",
            "out": m.get("p", ""),
        })
    return events


def rituais_to_events(rituais: list[dict]) -> list[dict]:
    """Converte rituais com data fixa em events. Cada ritual deve ter 'date' (ISO) e 'type'."""
    events = []
    for r in rituais:
        d = parse_date(r.get("date"))
        if not d:
            continue
        events.append({
            "d": d.strftime("%Y-%m-%d"),
            "type": r.get("type", "discovery"),
            "label": r.get("label", r.get("h3", "")),
            "who": r.get("who", ""),
            "dur": r.get("dur", ""),
            "out": r.get("out", ""),
        })
    return events


def expand_recurring(recurring: list[dict], period_start, period_end) -> list[dict]:
    """Expande rituais recorrentes em events.

    Cada item de `recurring` deve ter:
        type, label, freq (diario/semanal/quinzenal/mensal),
        start (ISO opcional, default = period_start),
        end (ISO opcional, default = period_end),
        weekday (0=seg ... 6=dom, opcional),
        who, dur, out
    """
    events = []
    if not (period_start and period_end):
        return events
    for r in recurring:
        freq = (r.get("freq") or "semanal").lower()
        delta_days = FREQ_DAYS.get(freq, 7)
        start = parse_date(r.get("start")) or period_start
        end = parse_date(r.get("end")) or period_end
        weekday = r.get("weekday")  # 0=monday
        cur = start
        # Ajusta para o weekday se especificado
        if weekday is not None:
            offset = (weekday - cur.weekday()) % 7
            cur = cur + dt.timedelta(days=offset)
        idx = 1
        while cur <= end:
            label = r.get("label", r.get("type", "evento"))
            if r.get("number_suffix"):
                label = f"{label} #{idx}"
            events.append({
                "d": cur.strftime("%Y-%m-%d"),
                "type": r.get("type", "checkin"),
                "label": label,
                "who": r.get("who", ""),
                "dur": r.get("dur", ""),
                "out": r.get("out", ""),
            })
            cur = cur + dt.timedelta(days=delta_days)
            idx += 1
    return events


def derive(data: dict, xlsx_path: Path) -> list[dict]:
    period_start = parse_date(data.get("period_start"))
    period_end = parse_date(data.get("period_end"))

    events: list[dict] = []
    events.extend(fases_from_xlsx(xlsx_path))

    roadmap = data.get("roadmap") or {}
    events.extend(milestones_to_events(roadmap.get("milestones") or []))

    com = data.get("comunicacao") or {}
    events.extend(rituais_to_events(com.get("rituais_calendar") or []))
    events.extend(expand_recurring(com.get("recurring_rituals") or [], period_start, period_end))

    # Sort by date
    events.sort(key=lambda e: e["d"])
    return events


def main() -> int:
    p = argparse.ArgumentParser(description="Deriva events[] do calendario a partir do xlsx + rituais.")
    p.add_argument("--input", required=True, help="Plan JSON (com data.comunicacao + data.roadmap)")
    p.add_argument("--xlsx", required=True, help="Path do Cronograma.xlsx")
    p.add_argument("--output", help="Path do events.json de saida (default: stdout)")
    p.add_argument("--inline-into", help="Caminho do plan.json para injetar events em data.events (modifica o file)")
    args = p.parse_args()

    try:
        data = json.loads(Path(args.input).read_text(encoding="utf-8"))
    except (FileNotFoundError, json.JSONDecodeError) as e:
        print(f"ERRO: nao consegui ler --input: {e}", file=sys.stderr)
        return 1

    xlsx = Path(args.xlsx)
    if not xlsx.exists():
        print(f"ERRO: --xlsx nao existe: {xlsx}", file=sys.stderr)
        return 1

    events = derive(data, xlsx)

    if args.inline_into:
        target = Path(args.inline_into)
        target_data = json.loads(target.read_text(encoding="utf-8"))
        target_data["events"] = events
        target.write_text(json.dumps(target_data, ensure_ascii=False, indent=2), encoding="utf-8")
        print(json.dumps({"ok": True, "events": len(events), "inlined_into": str(target)}))
        return 0

    if args.output:
        Path(args.output).write_text(json.dumps(events, ensure_ascii=False, indent=2), encoding="utf-8")
        print(json.dumps({"ok": True, "events": len(events), "output": args.output}))
    else:
        json.dump(events, sys.stdout, ensure_ascii=False, indent=2)
        sys.stdout.write("\n")

    return 0


if __name__ == "__main__":
    sys.exit(main())
