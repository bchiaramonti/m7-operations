#!/usr/bin/env python3
"""
render_html.py — Engine de renderizacao dos 10 HTMLs do plano de projeto.

Carrega templates de `templates/`, monta chunks dinamicos a partir do
project data + xlsx (cronograma) + events JSON (calendario), substitui
placeholders e escreve em `<project-dir>/1-planning/`.

Saida obrigatoria:
    <project-dir>/1-planning/plano-projeto.html
    <project-dir>/1-planning/artefatos/{contexto-escopo,eap,roadmap-marcos,
        okrs,recursos-dependencias,plano-comunicacao,riscos,cronograma,
        calendario}.html

Input: JSON via --input (ou stdin) com schema completo do plano.
Pre-requisito: --xlsx (Cronograma.xlsx ja gerado por generate_xlsx.py).

Uso:
    python3 render_html.py --input plan.json --xlsx 1-planning/Cronograma.xlsx \\
                          --output 1-planning/

Saida (stdout, JSON):
    {"ok": true, "files_written": [...], "warnings": [...]}
"""

import argparse
import datetime as dt
import json
import os
import re
import sys
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _lib import (  # noqa: E402
    LIME, OFF_WHITE, VERDE_CAQUI,
    MESES_BR, MESES_BR_FULL,
    build_footer_html, build_page_header_html, build_topbar_html,
    fmt_br, fmt_period, days_between,
    html_escape, load_logo_b64, load_template, parse_date,
    project_label, project_period_label, slugify,
)

try:
    import openpyxl
except ImportError:
    print("ERRO: openpyxl nao instalado. Execute: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

ARTIFACT_NAV = [
    ("01", "contexto-escopo.html",        "Contexto, Escopo & O que Não Faremos",  "Estrela guia, justificativa estratégica, fronteiras do projeto e exclusões explícitas."),
    ("02", "eap.html",                     "Estrutura Analítica do Projeto",        "WBS hierárquica em níveis com pacotes de trabalho e convenções."),
    ("03", "roadmap-marcos.html",          "Roadmap & Marcos",                       "Phase-bar, timeline visual e marcos com swim-lane por frente."),
    ("04", "okrs.html",                    "OKRs do Projeto",                        "Objetivos e resultados-chave mensuráveis com cadência de acompanhamento."),
    ("05", "recursos-dependencias.html",   "Recursos, Time & Dependências",          "Equipe, alocação por período e dependências com portfólio."),
    ("06", "plano-comunicacao.html",       "Plano de Comunicação",                    "Rituais de acompanhamento, RACI, canais e cadência de reporte."),
    ("07", "riscos.html",                  "Mapeamento & Matriz de Riscos",          "Heatmap probabilidade × impacto e contramedidas por risco."),
    ("08", "cronograma.html",              "Cronograma Detalhado",                    "Tabela WBS interativa com filtros (Fase / Ação / Etapa)."),
    ("09", "calendario.html",              "Calendário de Reuniões & Rituais",        "Visão mensal com eventos derivados do cronograma e rituais."),
]


# ---------------------------------------------------------------------------
# Helpers — string substitution
# ---------------------------------------------------------------------------

def substitute(template: str, subs: dict) -> str:
    """Aplica .replace() para cada (placeholder, valor) em subs."""
    out = template
    for k, v in subs.items():
        out = out.replace(k, str(v) if v is not None else "")
    return out


def check_remaining_placeholders(html: str, file_label: str) -> list[str]:
    """Detecta placeholders {{...}} nao substituidos. Retorna lista de warnings."""
    matches = re.findall(r"\{\{[a-zA-Z_][a-zA-Z0-9_]*\}\}", html)
    if matches:
        unique = sorted(set(matches))
        return [f"{file_label}: placeholders nao substituidos: {unique}"]
    return []


def write_html(output_path: Path, content: str, warnings: list[str], label: str) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(content, encoding="utf-8")
    warnings.extend(check_remaining_placeholders(content, label))


# ---------------------------------------------------------------------------
# Shared substitution helpers
# ---------------------------------------------------------------------------

def shared_subs(data: dict, logo_b64: str, num: str, h1: str,
                prev: tuple[str, str], next_: tuple[str, str]) -> dict:
    """
    Subs compartilhadas para os 9 artefatos (topbar, page-header, footer, project_label).
    `prev`/`next_` = (href, label).
    """
    plabel = project_label(data)
    period_label = project_period_label(data)
    return {
        "{{project_label}}": html_escape(plabel),
        "{{topbar_html}}": build_topbar_html(prev[0], prev[1], plabel, next_[0], next_[1]),
        "{{page_header_html}}": build_page_header_html(num, h1, logo_b64),
        "{{footer_html}}": build_footer_html(plabel, period_label),
    }


# ---------------------------------------------------------------------------
# Renderer 00 — landing (plano-projeto.html)
# ---------------------------------------------------------------------------

def render_landing(data: dict, logo_b64: str, output_dir: Path, warnings: list[str]) -> Path:
    template = load_template("plano-projeto.tmpl.html")

    nav_cards = []
    for num, href, title, desc in ARTIFACT_NAV:
        nav_cards.append(
            f'    <a href="artefatos/{href}" class="nav-card">\n'
            f'      <span class="card-num">{num}</span>\n'
            f'      <h3>{html_escape(title)}</h3>\n'
            f'      <p>{html_escape(desc)}</p>\n'
            f'      <span class="arrow">→</span>\n'
            f'    </a>'
        )

    period_start = data.get("period_start")
    period_end = data.get("period_end")
    dias = data.get("dias") or days_between(period_start, period_end)
    periodo = data.get("periodo") or fmt_period(period_start, period_end)
    project_code = (data.get("project_code") or "").strip()
    code_suffix = f" ({project_code})" if project_code else ""

    plabel = project_label(data)

    subs = {
        "{{logo_b64}}": logo_b64,
        "{{project_name}}": html_escape(data.get("project_name", "(sem nome)")),
        "{{project_code}}": html_escape(project_code or "—"),
        "{{project_code_suffix}}": html_escape(code_suffix),
        "{{project_subtitle}}": html_escape(data.get("project_subtitle", "")),
        "{{lider}}": html_escape(data.get("lider", "—")),
        "{{sponsor}}": html_escape(data.get("sponsor", "—")),
        "{{dias}}": str(dias),
        "{{periodo}}": html_escape(periodo),
        "{{estrela_guia}}": html_escape(data.get("estrela_guia", "")),
        "{{nav_cards_html}}": "\n".join(nav_cards),
        "{{footer_text}}": html_escape(
            f"M7 Investimentos — Equipe de Performance — {plabel} — {project_period_label(data) or ''}"
        ),
    }
    html = substitute(template, subs)
    output_path = output_dir / "plano-projeto.html"
    write_html(output_path, html, warnings, "plano-projeto.html")
    return output_path


# ---------------------------------------------------------------------------
# Renderer 01 — contexto-escopo.html
# ---------------------------------------------------------------------------

def render_contexto_escopo(data: dict, logo_b64: str, output_dir: Path,
                           warnings: list[str]) -> Path:
    template = load_template("artefatos/contexto-escopo.tmpl.html")
    ctx = data.get("contexto") or {}

    contexto_paragrafos = ctx.get("paragrafos_pre_quote") or []
    contexto_pos_quote = ctx.get("paragrafos_pos_quote") or []
    quote = ctx.get("quote") or {}
    pos_quote_h3 = ctx.get("pos_quote_h3") or ""

    # paragrafos_pre_quote / paragrafos_pos_quote: prosa livre, aceita HTML inline
    # (consistente com recursos.investimentos_paragrafos). Usar <strong>, <em>, <code>
    # diretamente no texto. Se precisar de `<` literal, escapar a mao na data.
    paragrafos_html = "\n".join(
        f'    <p>{p}</p>' for p in contexto_paragrafos
    )

    quote_html = ""
    if quote.get("text"):
        src = quote.get("source", "")
        quote_html = (
            '\n    <div class="quote-box">\n'
            f'      <p>"{html_escape(quote["text"])}"</p>\n'
            + (f'      <div class="source">{html_escape(src)}</div>\n' if src else "")
            + '    </div>'
        )

    pos_quote_html = ""
    if pos_quote_h3:
        pos_quote_html += f'\n    <h3>{html_escape(pos_quote_h3)}</h3>'
    pos_quote_html += "\n" + "\n".join(
        f'    <p>{p}</p>' for p in contexto_pos_quote
    )

    scope_yes = ctx.get("scope_yes") or []
    scope_no = ctx.get("scope_no") or []
    scope_yes_html = "\n".join(f'        <li>{html_escape(item)}</li>' for item in scope_yes)
    scope_no_html = "\n".join(f'        <li>{html_escape(item)}</li>' for item in scope_no)

    decisoes = ctx.get("decisoes") or []
    decisoes_html = "\n".join(
        f'        <tr><td>{html_escape(d.get("decisao",""))}</td>'
        f'<td>{html_escape(d.get("justificativa",""))}</td></tr>'
        for d in decisoes
    )

    conexoes = ctx.get("conexoes") or []
    conexoes_card_html = ""
    if conexoes:
        rows_html = "\n".join(
            f'        <tr>'
            f'<td>{html_escape(c.get("projeto",""))}</td>'
            f'<td><span class="badge badge-{html_escape(c.get("direcao_class","in"))}">'
            f'{html_escape(c.get("direcao_label","Entrada"))}</span></td>'
            f'<td>{html_escape(c.get("interface",""))}</td>'
            f'</tr>'
            for c in conexoes
        )
        conexoes_card_html = (
            '\n  <div class="card">\n'
            '    <h2>Conexões com Portfólio</h2>\n'
            '    <table class="dep-table">\n'
            '      <thead><tr><th>Projeto</th><th>Direção</th><th>Interface</th></tr></thead>\n'
            f'      <tbody>\n{rows_html}\n      </tbody>\n'
            '    </table>\n'
            '  </div>'
        )

    subs = shared_subs(
        data, logo_b64, "01", "Contexto, Escopo & O que Não Faremos",
        prev=("../plano-projeto.html", "Índice do Projeto"),
        next_=("eap.html", "Próximo: EAP"),
    )
    subs.update({
        "{{estrela_guia}}": html_escape(data.get("estrela_guia", "")),
        "{{contexto_paragrafos_html}}": paragrafos_html,
        "{{quote_box_html}}": quote_html,
        "{{contexto_pos_quote_html}}": pos_quote_html,
        "{{scope_yes_html}}": scope_yes_html,
        "{{scope_no_html}}": scope_no_html,
        "{{decisoes_html}}": decisoes_html,
        "{{conexoes_card_html}}": conexoes_card_html,
    })
    html = substitute(template, subs)
    output_path = output_dir / "artefatos" / "contexto-escopo.html"
    write_html(output_path, html, warnings, "contexto-escopo.html")
    return output_path


# ---------------------------------------------------------------------------
# Renderer 02 — eap.html (WBS tree)
# ---------------------------------------------------------------------------

def render_eap(data: dict, logo_b64: str, output_dir: Path, warnings: list[str]) -> Path:
    template = load_template("artefatos/eap.tmpl.html")
    eap = data.get("eap") or {}

    project_name = data.get("project_name", "(sem nome)")
    project_code = data.get("project_code", "")
    root_label = f"{project_code} {project_name}".strip() if project_code else project_name

    fases = eap.get("fases") or []
    nivel_4 = eap.get("nivel_4") or {}
    convencao = eap.get("convencao_html") or (
        "<strong>Convenção WBS:</strong> Nível 0 = Projeto. "
        "Nível 1 = Fases. Nível 2 = Subfases / Processos. Nível 3 = Pacotes de Trabalho."
    )

    # Build WBS tree HTML
    def render_node(node: dict, level: int) -> str:
        node_class = f"l{level}"
        if node.get("trans"):
            node_class += " trans"
        code = html_escape(node.get("code", ""))
        label = html_escape(node.get("label", ""))
        extra_html = ""
        if level == 0:
            extra = f'<div class="code">Nível 0</div><div class="label">{label}</div>'
            extra_html = extra
        elif level == 1:
            count = html_escape(node.get("count", ""))
            extra_html = (
                f'<div class="code">{code}</div>'
                f'<div class="label">{label}</div>'
                + (f'<div class="count">{count}</div>' if count else "")
            )
        elif level == 2:
            owner = html_escape(node.get("owner", ""))
            extra_html = (
                f'<div class="code">{code}</div>'
                f'<div class="label">{label}</div>'
                + (f'<div class="owner">{owner}</div>' if owner else "")
            )
        else:  # level 3+
            extra_html = (
                f'<div class="code">{code}</div>'
                f'<div class="label">{label}</div>'
            )
        children_html = ""
        children = node.get("children") or []
        if children:
            children_html = (
                '<ul>'
                + "".join(
                    f'<li>{render_node(c, level + 1)}</li>'
                    for c in children
                )
                + '</ul>'
            )
        return (
            f'<div class="node {node_class}">{extra_html}</div>'
            + children_html
        )

    root_node = {"label": root_label, "children": fases}
    wbs_tree_html = '<ul><li>' + render_node(root_node, 0) + '</li></ul>'

    # Nivel 4 table (optional)
    nivel_4_section_html = ""
    if nivel_4 and nivel_4.get("rows"):
        intro = nivel_4.get("intro_html", "")
        title = nivel_4.get("title", "Nível 4 — Pacotes de Trabalho por Processo")
        rows_html = "\n".join(
            f'      <tr>'
            f'<td class="wp-code">{html_escape(r.get("wbs",""))}</td>'
            f'<td>{html_escape(r.get("nome",""))}</td>'
            f'<td>{html_escape(r.get("nivel",""))}</td>'
            f'<td>{html_escape(r.get("descricao",""))}</td>'
            f'<td><span class="wp-fmt">{html_escape(r.get("formato",""))}</span></td>'
            f'</tr>'
            for r in nivel_4["rows"]
        )
        nivel_4_section_html = (
            f'\n  <div class="section-title">{html_escape(title)}</div>\n'
            + (f'  <p style="font-size:13px;color:var(--cinza-400);margin-bottom:16px;">{intro}</p>\n' if intro else "")
            + '  <table class="wp-table">\n'
            '    <thead><tr><th>WBS</th><th>Pacote de Trabalho</th><th>Nível</th><th>Descrição</th><th>Formato</th></tr></thead>\n'
            f'    <tbody>\n{rows_html}\n    </tbody>\n'
            '  </table>'
        )

    subs = shared_subs(
        data, logo_b64, "02", "Estrutura Analítica do Projeto (EAP)",
        prev=("contexto-escopo.html", "Contexto & Escopo"),
        next_=("roadmap-marcos.html", "Próximo: Roadmap"),
    )
    subs.update({
        "{{wbs_tree_html}}": wbs_tree_html,
        "{{nivel_4_section_html}}": nivel_4_section_html,
        "{{convencao_wbs_html}}": convencao,
    })
    html = substitute(template, subs)
    output_path = output_dir / "artefatos" / "eap.html"
    write_html(output_path, html, warnings, "eap.html")
    return output_path


# ---------------------------------------------------------------------------
# Renderer 03 — roadmap-marcos.html
# ---------------------------------------------------------------------------

def render_roadmap_marcos(data: dict, logo_b64: str, output_dir: Path,
                          warnings: list[str]) -> Path:
    template = load_template("artefatos/roadmap-marcos.tmpl.html")
    roadmap = data.get("roadmap") or {}

    # Month columns
    period_start = parse_date(data.get("period_start"))
    period_end = parse_date(data.get("period_end"))
    months: list[tuple[int, int]] = []
    if period_start and period_end:
        cur = dt.datetime(period_start.year, period_start.month, 1)
        end_anchor = dt.datetime(period_end.year, period_end.month, 1)
        while cur <= end_anchor:
            months.append((cur.year, cur.month))
            if cur.month == 12:
                cur = dt.datetime(cur.year + 1, 1, 1)
            else:
                cur = dt.datetime(cur.year, cur.month + 1, 1)
    n_months = max(len(months), 1)

    # Colunas ponderadas pelos dias que cada mês contribui ao período
    fr_cols = _compute_month_fractions(period_start, period_end, months)

    months_row_html = "\n".join(
        f'      <div class="month-cell">{MESES_BR[m].upper()}<span class="year">{y}</span></div>'
        for (y, m) in months
    )

    # Data lanes — agrupadas por `phase` (campo opcional no schema).
    # Se alguma lane tiver `phase`, emite phase-dividers entre os blocos.
    lanes = roadmap.get("lanes") or []
    lanes_html = _render_lanes_with_phase_dividers(lanes, period_start, period_end)

    # Milestones — lane do topo + tabela abaixo (day-linear scale)
    milestones = roadmap.get("milestones") or []
    milestones_lane_html = _render_milestones_lane(milestones, period_start, period_end)
    marcos_table_rows_html = "\n".join(
        _render_marco_row(m, idx) for idx, m in enumerate(milestones)
    )

    subs = shared_subs(
        data, logo_b64, "03", "Roadmap & Marcos",
        prev=("eap.html", "EAP"),
        next_=("okrs.html", "Próximo: OKRs"),
    )
    subs.update({
        "{{n_months}}": str(n_months),
        "{{fr_cols}}": fr_cols,
        "{{months_row_html}}": months_row_html,
        "{{milestones_lane_html}}": milestones_lane_html,
        "{{lanes_html}}": lanes_html,
        "{{marcos_table_rows_html}}": marcos_table_rows_html,
    })
    html = substitute(template, subs)
    output_path = output_dir / "artefatos" / "roadmap-marcos.html"
    write_html(output_path, html, warnings, "roadmap-marcos.html")
    return output_path


def _compute_month_fractions(period_start, period_end,
                             months: list[tuple[int, int]]) -> str:
    """Gera `grid-template-columns` ponderado pelos dias reais que cada mês
    contribui ao período. Retorna algo como '5fr 30fr 31fr 30fr 18fr'.

    Quando o período não começa no dia 1 ou não termina no último dia do mês,
    isso mantém as colunas do header alinhadas com as bars/ticks (day-linear).
    """
    if not (period_start and period_end and months):
        return " ".join("1fr" for _ in months) if months else "1fr"
    parts: list[str] = []
    for (y, m) in months:
        first = dt.date(y, m, 1)
        next_first = dt.date(y + 1, 1, 1) if m == 12 else dt.date(y, m + 1, 1)
        last = next_first - dt.timedelta(days=1)
        seg_start = max(first, period_start.date() if hasattr(period_start, "date") else period_start)
        seg_end = min(last, period_end.date() if hasattr(period_end, "date") else period_end)
        days = (seg_end - seg_start).days + 1 if seg_end >= seg_start else 1
        parts.append(f"{days}fr")
    return " ".join(parts)


def _render_milestones_lane(milestones: list[dict],
                            period_start, period_end) -> str:
    """Lane no topo do roadmap com ticks alternados (top/bottom) ao redor de um trilho central.
    Usa posicionamento day-linear (mesma escala das bars e QRs) para manter alinhamento
    visual entre os marcos do topo e as entregas das frentes."""
    if not milestones:
        return ""

    ticks_html_parts = []
    for idx, m in enumerate(milestones):
        m_date = parse_date(m.get("date_iso") or m.get("date"))
        if not m_date:
            continue
        left = _percent_anchor(m_date, period_start, period_end)
        is_major = m.get("major", False)
        gate_class = " gate" if is_major else " regular"
        position = "top" if idx % 2 == 0 else "bottom"
        lbl = html_escape(m.get("lbl") or _short_lbl(m.get("h4", "")))
        date_str = html_escape(m.get("date") or fmt_br(m_date))
        desc = html_escape(m.get("desc") or "")

        if position == "top":
            ticks_html_parts.append(
                f'        <div class="tick top{gate_class}" style="left:{left:.2f}%;">\n'
                f'          <div class="lbl">{lbl}</div>\n'
                f'          <div class="date">{date_str}</div>\n'
                + (f'          <div class="desc">{desc}</div>\n' if desc else "")
                + '          <div class="connector"></div>\n'
                '          <div class="dot"></div>\n'
                '        </div>'
            )
        else:
            ticks_html_parts.append(
                f'        <div class="tick bottom{gate_class}" style="left:{left:.2f}%;">\n'
                '          <div class="dot"></div>\n'
                '          <div class="connector"></div>\n'
                f'          <div class="lbl">{lbl}</div>\n'
                f'          <div class="date">{date_str}</div>\n'
                + (f'          <div class="desc">{desc}</div>\n' if desc else "")
                + '        </div>'
            )

    return (
        '    <div class="lane milestones">\n'
        '      <div class="lane-label">\n'
        '        <div class="code">M</div>\n'
        '        <div class="name">Marcos do Projeto</div>\n'
        f'        <div class="owner">{len(milestones)} gates</div>\n'
        '      </div>\n'
        '      <div class="track">\n'
        '        <div class="rail"></div>\n'
        + "\n".join(ticks_html_parts) + "\n"
        + '      </div>\n'
        + '    </div>'
    )


def _render_marco_row(m: dict, idx: int) -> str:
    """Linha da Tabela de Marcos (Tipo · Marco · Data · WBS · Descrição)."""
    alt_class = " alt" if idx % 2 == 1 else ""
    tipo_class = " gate" if m.get("major") else ""
    wbs = html_escape(m.get("wbs", ""))
    wbs_html = f'<span class="chip">{wbs}</span>' if wbs else ""
    return (
        f'      <div class="data-row{alt_class}">\n'
        f'        <div class="col-tipo"><span class="tipo-dot{tipo_class}"></span></div>\n'
        f'        <div class="col-marco">{html_escape(m.get("h4",""))}</div>\n'
        f'        <div class="col-data">{html_escape(m.get("date",""))}</div>\n'
        f'        <div class="col-wbs">{wbs_html}</div>\n'
        f'        <div class="col-desc">{html_escape(m.get("p",""))}</div>\n'
        '      </div>'
    )


def _short_lbl(h4: str) -> str:
    """Fallback para tick label: pega o trecho antes do primeiro ' · ' ou trunca em 18 chars."""
    if not h4:
        return ""
    if " · " in h4:
        return h4.split(" · ", 1)[0].upper()
    return h4[:18].upper()


# Constantes visuais do swim-lane (espelham as CSS vars no template)
_BAR_H_PX = 46
_ROW_GAP_PX = 8
_ROW_STEP_PX = _BAR_H_PX + _ROW_GAP_PX  # 54
_BAR_TOP_BASE_PX = 8
_LANE_PAD_BOTTOM_PX = 8
_OVERLAP_TOL_PCT = 0.05

# Largura estimada do track (px) — usada para decidir se título cabe dentro da bar.
# Corresponde a container 1440 - 2*32 padding - 180 lane-label - borders ≈ 1195.
_TRACK_WIDTH_PX = 1195
_CHAR_W_PX = 6.2       # largura média de char no font-size 10.5 (com margem p/ acentos)
_SAFETY_CHARS = 2      # margem conservadora p/ evitar truncate
_BAR_LINE_CLAMP = 2    # -webkit-line-clamp:2 permite título em até 2 linhas


def _assign_rows(bars_pos: list[tuple[float, float]]) -> list[int]:
    """Greedy: atribui cada bar a uma row sem overlap temporal (com tolerância
    _OVERLAP_TOL_PCT). Retorna lista de row-index na mesma ordem das bars."""
    if not bars_pos:
        return []
    # Rows: cada row armazena o maior `right` (left+width) já alocado
    rows: list[float] = []
    assignment: list[int] = []
    # Mantém ordem original, mas prioriza bars que começam mais cedo
    order = sorted(range(len(bars_pos)), key=lambda i: bars_pos[i][0])
    result: list[int] = [0] * len(bars_pos)
    for i in order:
        left, width = bars_pos[i]
        right = left + width
        placed = False
        for r_idx, r_right in enumerate(rows):
            if left + _OVERLAP_TOL_PCT >= r_right:
                rows[r_idx] = right
                result[i] = r_idx
                placed = True
                break
        if not placed:
            rows.append(right)
            result[i] = len(rows) - 1
    return result


def _chars_fit_in_bar(width_pct: float) -> int:
    """Estima quantos chars cabem dentro de uma bar de `width_pct` do track.
    Considera que o CSS aplica `-webkit-line-clamp:2` (até 2 linhas) e aplica
    _SAFETY_CHARS de margem para evitar truncate inesperado."""
    usable_px = (width_pct / 100.0) * _TRACK_WIDTH_PX - 28  # 28 = padding 6+18+margem
    if usable_px <= 0:
        return 0
    per_line = max(0, usable_px / _CHAR_W_PX - _SAFETY_CHARS)
    return int(per_line * _BAR_LINE_CLAMP)


def _format_br_short(d) -> str:
    """Formato curto para macro-flotilha: '28/abr' (sem ano)."""
    parsed = parse_date(d)
    if not parsed:
        return ""
    return f"{parsed.day:02d}/{MESES_BR[parsed.month]}"


def _render_lanes_with_phase_dividers(lanes: list[dict], period_start, period_end) -> str:
    """Renderiza todas as lanes, injetando phase-dividers entre blocos de fases
    quando pelo menos uma lane tem campo `phase`. Mantém a ordem original das lanes."""
    if not lanes:
        return ""
    has_any_phase = any(lane.get("phase") for lane in lanes)

    # Conta lanes por fase (preservando ordem de aparição)
    phase_order: list[str] = []
    phase_counts: dict[str, int] = {}
    if has_any_phase:
        for lane in lanes:
            p = lane.get("phase") or ""
            if not p:
                continue
            if p not in phase_counts:
                phase_order.append(p)
                phase_counts[p] = 0
            phase_counts[p] += 1

    parts: list[str] = []
    emitted_phase: set[str] = set()
    for lane in lanes:
        phase = lane.get("phase") or ""
        if has_any_phase and phase and phase not in emitted_phase:
            emitted_phase.add(phase)
            count = phase_counts[phase]
            parts.append(_render_phase_divider(phase, count))
        parts.append(_render_lane(lane, period_start, period_end))
    return "\n".join(parts)


def _render_phase_divider(phase: str, count: int) -> str:
    """Banner clicável entre blocos de fases, com contador de lanes e seta."""
    plural = "s" if count != 1 else ""
    return (
        f'    <div class="lane phase-divider" data-phase="{html_escape(phase)}" '
        'role="button" tabindex="0" aria-expanded="true" '
        'title="Clique para recolher/expandir toda a fase">\n'
        f'      <div class="phase-title">Fase · {html_escape(phase)}'
        f'<span class="phase-meta">{count} lane{plural}</span>'
        '<span class="phase-toggle" aria-hidden="true">&#9662;</span></div>\n'
        '    </div>'
    )


def _render_lane(lane: dict, period_start, period_end) -> str:
    """Renderiza uma .lane do swim-lane com bars (com stacking + bar-ext-label para
    estreitas), qrs de governança e macro-flotilha (visível quando recolhida)."""
    is_gov = lane.get("is_gov", False)
    code = html_escape(lane.get("code", ""))
    name = html_escape(lane.get("name", ""))
    owner = html_escape(lane.get("owner", ""))
    phase = lane.get("phase") or ""

    # ── 1. Computa posições de bars e determina rows (stacking)
    bars = lane.get("bars") or []
    bar_positions: list[tuple[float, float]] = []
    for bar in bars:
        bar_start = parse_date(bar.get("start"))
        bar_end = parse_date(bar.get("end"))
        bar_positions.append(_percent_position(bar_start, bar_end, period_start, period_end))
    rows = _assign_rows(bar_positions)
    max_rows = (max(rows) + 1) if rows else 0

    # ── 2. Renderiza cada bar + ext-label quando necessário
    bars_html_parts: list[str] = []
    ext_labels_html_parts: list[str] = []
    for idx, bar in enumerate(bars):
        left, width = bar_positions[idx]
        row = rows[idx]
        top_px = _BAR_TOP_BASE_PX + row * _ROW_STEP_PX
        bar_class = bar.get("class", "v-dark")
        title = bar.get("title", "")
        title_escaped = html_escape(title)

        # Narrow detection: título não cabe dentro da bar?
        max_chars = _chars_fit_in_bar(width)
        is_narrow = len(title) > max_chars
        extra_cls = " narrow" if is_narrow else ""

        bars_html_parts.append(
            f'        <div class="bar {html_escape(bar_class)}{extra_cls}" '
            f'style="left:{left:.1f}%; width:{width:.1f}%; top:{top_px}px;">'
            f'<span class="title">{title_escaped}</span></div>'
        )

        if is_narrow:
            # Se a bar termina após 70% do track, flip o label para a esquerda
            right_pct = left + width
            if right_pct > 70.0:
                ext_labels_html_parts.append(
                    f'        <div class="bar-ext-label flip-left" '
                    f'style="right:{100.0 - left:.2f}%; top:{top_px}px;">'
                    f'<span class="t">{title_escaped}</span></div>'
                )
            else:
                ext_labels_html_parts.append(
                    f'        <div class="bar-ext-label" '
                    f'style="left:{right_pct:.2f}%; top:{top_px}px;">'
                    f'<span class="t">{title_escaped}</span></div>'
                )

    # ── 3. Qr badges (governance lane) — day-linear
    qrs = lane.get("qrs") or []
    qrs_html_parts: list[str] = []
    for qr in qrs:
        qr_date = parse_date(qr.get("date"))
        left = _percent_anchor(qr_date, period_start, period_end)
        label = html_escape(qr.get("label", ""))
        qrs_html_parts.append(
            f'        <div class="qr" style="left:{left:.1f}%;">'
            f'<span class="qr-label">{label}</span></div>'
        )

    # ── 4. Macro-flotilha: span da lane (min start, max end das bars + qrs)
    macro_html = _render_lane_macro(lane, bars, qrs, period_start, period_end)

    # ── 5. Altura dinâmica do track (acomoda max_rows sem cortar a última bar)
    if max_rows > 0:
        track_min_h = _BAR_TOP_BASE_PX + max_rows * _ROW_STEP_PX + _LANE_PAD_BOTTOM_PX - _ROW_GAP_PX
        track_style = f' style="min-height:{track_min_h}px;"'
    else:
        track_style = ' style="min-height:62px;"' if not is_gov else ""

    # ── 6. Monta lane-label (com toggle button) e track
    lane_class = "lane gov" if is_gov else "lane"
    data_attrs = (
        f' data-phase="{html_escape(phase)}"' if phase else ""
    ) + (f' data-code="{code}"' if code else "") + ' aria-expanded="true"'
    toggle_btn = (
        f'<button type="button" class="lane-toggle" '
        f'aria-label="Recolher/expandir {code}" title="Recolher/expandir">&#9662;</button>'
    )
    lane_label_html = (
        '      <div class="lane-label">'
        f'{toggle_btn}\n'
        f'        <div class="code">{code}</div>\n'
        f'        <div class="name">{name}</div>\n'
        + (f'        <div class="owner">{owner}</div>\n' if owner else "")
        + '      </div>'
    )
    track_inner_parts = []
    if macro_html:
        track_inner_parts.append(macro_html)
    track_inner_parts.extend(bars_html_parts)
    track_inner_parts.extend(ext_labels_html_parts)
    track_inner_parts.extend(qrs_html_parts)
    track_html = (
        f'      <div class="track"{track_style}>\n'
        + "\n".join(track_inner_parts) + "\n"
        + '      </div>'
    )
    return (
        f'    <div class="{lane_class}"{data_attrs}>\n'
        + lane_label_html + "\n"
        + track_html + "\n"
        + '    </div>'
    )


def _render_lane_macro(lane: dict, bars: list[dict], qrs: list[dict],
                       period_start, period_end) -> str:
    """Macro-flotilha da lane: um chip lime cobrindo min(start)..max(end) das
    bars + qrs. Renderizada sempre; CSS a esconde por padrão e a revela quando
    a própria lane está .collapsed."""
    dates_start: list = []
    dates_end: list = []
    for bar in bars:
        bs = parse_date(bar.get("start"))
        be = parse_date(bar.get("end"))
        if bs: dates_start.append(bs)
        if be: dates_end.append(be)
    for qr in qrs:
        d = parse_date(qr.get("date"))
        if d:
            dates_start.append(d)
            dates_end.append(d)
    if not (dates_start and dates_end and period_start and period_end):
        return ""
    lane_min = min(dates_start)
    lane_max = max(dates_end)
    left, width = _percent_position(lane_min, lane_max, period_start, period_end)
    label = html_escape(lane.get("name", ""))
    start_txt = _format_br_short(lane_min)
    end_txt = _format_br_short(lane_max)
    return (
        f'        <div class="lane-macro-bar" '
        f'style="left:{left:.3f}%; width:{width:.3f}%;" aria-hidden="true">'
        f'<span class="lmb-start">{start_txt}</span>'
        f'<span class="lmb-label">{label}</span>'
        f'<span class="lmb-end">{end_txt}</span></div>'
    )


def _percent_position(start, end, period_start, period_end) -> tuple[float, float]:
    """Calcula left% e width% de uma barra dentro do periodo total."""
    if not (start and end and period_start and period_end):
        return (0.0, 100.0)
    total = (period_end - period_start).days or 1
    left_d = (start - period_start).days
    width_d = (end - start).days + 1
    left_pct = max(0.0, min(100.0, 100.0 * left_d / total))
    width_pct = max(0.5, min(100.0 - left_pct, 100.0 * width_d / total))
    return (left_pct, width_pct)


def _percent_anchor(point, period_start, period_end) -> float:
    """Calcula left% de um ponto (marco) dentro do periodo total."""
    if not (point and period_start and period_end):
        return 0.0
    total = (period_end - period_start).days or 1
    d = (point - period_start).days
    return max(0.0, min(100.0, 100.0 * d / total))


# ---------------------------------------------------------------------------
# Renderer 04 — okrs.html
# ---------------------------------------------------------------------------

def render_okrs(data: dict, logo_b64: str, output_dir: Path,
                warnings: list[str]) -> Path:
    template = load_template("artefatos/okrs.tmpl.html")
    okrs = data.get("okrs") or []
    cadence_items = data.get("okrs_cadence") or []

    okr_blocks_html = []
    for i, obj in enumerate(okrs, start=1):
        krs_html_parts = []
        for kr in obj.get("krs") or []:
            metric_html = kr.get("metric", "")
            krs_html_parts.append(
                '      <div class="kr">\n'
                f'        <div class="kr-num">{html_escape(kr.get("num",""))}</div>\n'
                '        <div class="kr-content">\n'
                f'          <h4>{html_escape(kr.get("h4",""))}</h4>\n'
                f'          <div class="metric">{metric_html}</div>\n'
                '        </div>\n'
                f'        <div class="kr-target">{html_escape(kr.get("target",""))}'
                + (f'<span class="unit">{html_escape(kr.get("unit",""))}</span>' if kr.get("unit") else "")
                + '</div>\n'
                '      </div>'
            )

        okr_blocks_html.append(
            '  <div class="okr-block">\n'
            '    <div class="obj-card">\n'
            f'      <div class="obj-num">Objetivo {obj.get("num", i)}</div>\n'
            f'      <h2>{html_escape(obj.get("descricao", ""))}</h2>\n'
            '    </div>\n'
            '    <div class="kr-list">\n'
            + "\n".join(krs_html_parts) + "\n"
            '    </div>\n'
            '  </div>'
        )

    cadence_items_html = "\n".join(
        '      <div class="cadence-item">\n'
        f'        <div class="freq">{html_escape(c.get("freq",""))}</div>\n'
        f'        <div class="desc">{html_escape(c.get("desc",""))}</div>\n'
        '      </div>'
        for c in cadence_items
    )

    subs = shared_subs(
        data, logo_b64, "04", "OKRs do Projeto",
        prev=("roadmap-marcos.html", "Roadmap"),
        next_=("recursos-dependencias.html", "Próximo: Recursos"),
    )
    subs.update({
        "{{okr_blocks_html}}": "\n".join(okr_blocks_html),
        "{{cadence_items_html}}": cadence_items_html,
    })
    html = substitute(template, subs)
    output_path = output_dir / "artefatos" / "okrs.html"
    write_html(output_path, html, warnings, "okrs.html")
    return output_path


# ---------------------------------------------------------------------------
# Renderer 05 — recursos-dependencias.html
# ---------------------------------------------------------------------------

def render_recursos(data: dict, logo_b64: str, output_dir: Path,
                    warnings: list[str]) -> Path:
    template = load_template("artefatos/recursos-dependencias.tmpl.html")
    recursos = data.get("recursos") or {}

    team = recursos.get("team") or []
    team_cards_html = "\n".join(_render_team_card(member) for member in team)

    alloc = recursos.get("alloc") or {}
    periods = alloc.get("periods") or []
    period_headers_html = "\n".join(
        f'          <th>{html_escape(p)}</th>' for p in periods
    )
    alloc_rows_html = "\n".join(
        _render_alloc_row(row, periods) for row in (alloc.get("rows") or [])
    )

    deps = recursos.get("dependencias") or []
    deps_rows_html = "\n".join(
        f'      <tr>'
        f'<td><strong>{html_escape(d.get("projeto",""))}</strong></td>'
        f'<td><span class="badge badge-{html_escape(d.get("tipo_class","in"))}">{html_escape(d.get("tipo_label","Entrada"))}</span></td>'
        f'<td>{html_escape(d.get("interface",""))}</td>'
        f'<td>{html_escape(d.get("risco",""))}</td>'
        f'</tr>'
        for d in deps
    )

    invest_paragraphs = recursos.get("investimentos_paragrafos") or []
    invest_html = "\n".join(f'    <p>{p}</p>' for p in invest_paragraphs)

    subs = shared_subs(
        data, logo_b64, "05", "Recursos, Time & Dependências",
        prev=("okrs.html", "OKRs"),
        next_=("plano-comunicacao.html", "Próximo: Comunicação"),
    )
    subs.update({
        "{{team_cards_html}}": team_cards_html,
        "{{alloc_period_headers_html}}": period_headers_html,
        "{{alloc_rows_html}}": alloc_rows_html,
        "{{dependencias_rows_html}}": deps_rows_html,
        "{{investimentos_html}}": invest_html,
    })
    html = substitute(template, subs)
    output_path = output_dir / "artefatos" / "recursos-dependencias.html"
    write_html(output_path, html, warnings, "recursos-dependencias.html")
    return output_path


def _render_team_card(member: dict) -> str:
    is_lead = member.get("lead", False)
    role = member.get("role", "Membro")
    role_class = "badge-lider" if is_lead else "badge-dono"
    blocks_html = "".join(
        f'<span class="block-chip{" trans" if b.get("trans") else ""}">{html_escape(b.get("label",""))}</span>'
        for b in (member.get("blocks") or [])
    )
    return (
        f'    <div class="team-card{" lead" if is_lead else ""}">\n'
        f'      <span class="role-badge {role_class}">{html_escape(role)}</span>\n'
        f'      <h3>{html_escape(member.get("name",""))}</h3>\n'
        f'      <div class="area">{html_escape(member.get("area",""))}</div>\n'
        + (f'      <div class="blocks">{blocks_html}</div>\n' if blocks_html else "")
        + f'      <div class="alloc">{html_escape(member.get("alloc",""))}</div>\n'
        + '    </div>'
    )


def _render_alloc_row(row: dict, periods: list[str]) -> str:
    name = html_escape(row.get("name", ""))
    cells = row.get("cells") or {}
    cell_html_parts = []
    for p in periods:
        cell = cells.get(p)
        if cell:
            label = html_escape(cell.get("label", ""))
            cls = html_escape(cell.get("class", "active"))
            cell_html_parts.append(f'<td class="cell-{cls}">{label}</td>')
        else:
            cell_html_parts.append('<td></td>')
    return f'      <tr><td>{name}</td>{"".join(cell_html_parts)}</tr>'


# ---------------------------------------------------------------------------
# Renderer 06 — plano-comunicacao.html
# ---------------------------------------------------------------------------

def render_plano_comunicacao(data: dict, logo_b64: str, output_dir: Path,
                             warnings: list[str]) -> Path:
    template = load_template("artefatos/plano-comunicacao.tmpl.html")
    com = data.get("comunicacao") or {}

    rituais = com.get("rituais") or []
    ritual_cards_html = "\n".join(_render_ritual_card(r) for r in rituais)

    raci = com.get("raci") or {}
    papeis = raci.get("papeis") or []
    raci_papel_headers_html = "\n".join(
        f'          <th>{html_escape(p)}</th>' for p in papeis
    )
    raci_rows_html = "\n".join(
        _render_raci_row(row, papeis) for row in (raci.get("rows") or [])
    )

    channels = com.get("channels") or []
    channels_html = "\n".join(
        '    <div class="channel">\n'
        f'      <div class="icon">{html_escape(c.get("icon",""))}</div>\n'
        f'      <h4>{html_escape(c.get("h4",""))}</h4>\n'
        f'      <p>{html_escape(c.get("p",""))}</p>\n'
        '    </div>'
        for c in channels
    )

    subs = shared_subs(
        data, logo_b64, "06", "Plano de Comunicação",
        prev=("recursos-dependencias.html", "Recursos"),
        next_=("riscos.html", "Próximo: Riscos"),
    )
    subs.update({
        "{{ritual_cards_html}}": ritual_cards_html,
        "{{raci_papel_headers_html}}": raci_papel_headers_html,
        "{{raci_rows_html}}": raci_rows_html,
        "{{channels_html}}": channels_html,
    })
    html = substitute(template, subs)
    output_path = output_dir / "artefatos" / "plano-comunicacao.html"
    write_html(output_path, html, warnings, "plano-comunicacao.html")
    return output_path


def _render_ritual_card(r: dict) -> str:
    cls = " highlight" if r.get("highlight") else ""
    meta = r.get("meta") or []
    meta_html = "".join(
        f'<span>{html_escape(m.get("label",""))}: <strong>{html_escape(m.get("value",""))}</strong></span>'
        for m in meta
    )
    return (
        f'    <div class="ritual-card{cls}">\n'
        f'      <div class="freq">{html_escape(r.get("freq",""))}</div>\n'
        f'      <h3>{html_escape(r.get("h3",""))}</h3>\n'
        f'      <p>{html_escape(r.get("p",""))}</p>\n'
        + (f'      <div class="meta">{meta_html}</div>\n' if meta_html else "")
        + '    </div>'
    )


def _render_raci_row(row: dict, papeis: list[str]) -> str:
    atividade = html_escape(row.get("atividade", ""))
    atribuicoes = row.get("atribuicoes") or {}
    cells = []
    for p in papeis:
        v = atribuicoes.get(p, "").upper().strip()
        if v in ("R", "A", "C", "I"):
            cls = f"raci-{v.lower()}"
            cells.append(f'<td><span class="{cls}">{v}</span></td>')
        else:
            cells.append('<td></td>')
    return f'      <tr><td>{atividade}</td>{"".join(cells)}</tr>'


# ---------------------------------------------------------------------------
# Renderer 07 — riscos.html
# ---------------------------------------------------------------------------

def render_riscos(data: dict, logo_b64: str, output_dir: Path,
                  warnings: list[str]) -> Path:
    template = load_template("artefatos/riscos.tmpl.html")
    riscos = data.get("riscos") or {}
    items = riscos.get("items") or []

    # Heatmap: 3x3 grid (prob: alta=top, baixa=bottom; impacto: baixo=left, alto=right)
    # Coordenadas: (prob, imp) -> ["high","med","low"] x ["low","med","high"]
    grid = {(p, i): [] for p in ("high", "med", "low") for i in ("low", "med", "high")}
    for it in items:
        p = (it.get("prob") or "med").lower()
        i = (it.get("imp") or "med").lower()
        if p in ("alta", "high"): p_key = "high"
        elif p in ("baixa", "low"): p_key = "low"
        else: p_key = "med"
        if i in ("alto", "high"): i_key = "high"
        elif i in ("baixo", "low"): i_key = "low"
        else: i_key = "med"
        grid[(p_key, i_key)].append(it.get("id", ""))

    cell_order = [
        ("high", "low"), ("high", "med"), ("high", "high"),
        ("med", "low"),  ("med", "med"),  ("med", "high"),
        ("low", "low"),  ("low", "med"),  ("low", "high"),
    ]
    heatmap_cells_html = "\n".join(
        _render_heatmap_cell(p, i, grid[(p, i)]) for (p, i) in cell_order
    )

    # Risk legend items
    risk_items_html = "\n".join(_render_risk_item(it) for it in items)

    # Contramedidas
    contramedidas = riscos.get("contramedidas") or items
    contramedidas_rows_html = "\n".join(
        f'      <tr>\n'
        f'        <td><strong>{html_escape(c.get("id",""))}</strong></td>\n'
        f'        <td>{html_escape(c.get("risco","") or c.get("h4",""))}</td>\n'
        f'        <td><span class="sev sev-{_sev_class(c.get("severity"))}">'
        f'{html_escape(_sev_label(c.get("severity")))}</span></td>\n'
        f'        <td>{html_escape(c.get("contramedida",""))}</td>\n'
        f'        <td>{html_escape(c.get("acao_mitigacao", c.get("mitigation","")))}</td>\n'
        f'        <td>{html_escape(c.get("trigger",""))}</td>\n'
        f'      </tr>'
        for c in contramedidas
    )

    subs = shared_subs(
        data, logo_b64, "07", "Mapeamento & Matriz de Riscos",
        prev=("plano-comunicacao.html", "Comunicação"),
        next_=("cronograma.html", "Próximo: Cronograma"),
    )
    subs.update({
        "{{heatmap_cells_html}}": heatmap_cells_html,
        "{{risk_items_html}}": risk_items_html,
        "{{contramedidas_rows_html}}": contramedidas_rows_html,
    })
    html = substitute(template, subs)
    output_path = output_dir / "artefatos" / "riscos.html"
    write_html(output_path, html, warnings, "riscos.html")
    return output_path


def _render_heatmap_cell(prob: str, imp: str, ids: list[str]) -> str:
    cls = f"heat-{prob}-{imp}"
    dots = "".join(
        f'<span class="risk-dot">{html_escape(rid)}</span>' for rid in ids
    )
    inner = f'<div class="risk-dots">{dots}</div>' if dots else ""
    return f'          <div class="heat-cell {cls}">{inner}</div>'


def _render_risk_item(it: dict) -> str:
    sev = _sev_class(it.get("severity"))
    sev_id_class = "crit" if sev == "crit" else sev
    mitigation = it.get("mitigation", "")
    mitigation_html = (
        '          <div class="mitigation">\n'
        '            <strong>Ação de Mitigação</strong>\n'
        f'            <span>{html_escape(mitigation)}</span>\n'
        '          </div>'
        if mitigation else ""
    )
    return (
        '      <div class="risk-item">\n'
        f'        <div class="risk-id {sev_id_class}">{html_escape(it.get("id",""))}</div>\n'
        '        <div class="risk-content">\n'
        f'          <h4>{html_escape(it.get("h4",""))}</h4>\n'
        '          <div class="tags">'
        f'<span class="risk-tag tag-prob">Prob: {html_escape(it.get("prob",""))}</span>'
        f'<span class="risk-tag tag-imp">Imp: {html_escape(it.get("imp",""))}</span>'
        '</div>\n'
        f'          <p>{html_escape(it.get("p",""))}</p>\n'
        + mitigation_html + "\n"
        + '        </div>\n'
        '      </div>'
    )


def _sev_class(severity) -> str:
    s = (severity or "med").lower()
    if s in ("crit", "critico", "critical", "crítico"): return "crit"
    if s in ("high", "alto"): return "high"
    return "med"


def _sev_label(severity) -> str:
    s = _sev_class(severity)
    return {"crit": "Crítico", "high": "Alto", "med": "Médio"}[s]


# ---------------------------------------------------------------------------
# Renderer 08 — cronograma.html (le do Cronograma.xlsx)
# ---------------------------------------------------------------------------

def render_cronograma(data: dict, logo_b64: str, xlsx_path: Path,
                      output_dir: Path, warnings: list[str]) -> Path:
    template = load_template("artefatos/cronograma.tmpl.html")
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Cronograma.xlsx nao existe em: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    ws = wb.active
    HEADER_ROW = 4
    FIRST_DATA_ROW = 5
    # detect last data row
    last_row = HEADER_ROW
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        if ws.cell(r, 2).value:
            last_row = r

    rows_html_parts = []
    stats = {"fases": 0, "acoes": 0, "etapas": 0}
    last_close_phase_row = None  # para marcar a ultima Fase como close (lime)

    raw_rows = []
    for r in range(FIRST_DATA_ROW, last_row + 1):
        no = str(ws.cell(r, 2).value or "").strip()
        if not no:
            continue
        tipo = str(ws.cell(r, 3).value or "").strip()
        etapa = ws.cell(r, 4).value or ""
        responsavel = ws.cell(r, 5).value or ""
        ini = ws.cell(r, 6).value
        fim = ws.cell(r, 7).value
        entreg = ws.cell(r, 11).value or ""

        ini_str = fmt_br(ini)
        fim_str = fmt_br(fim)

        raw_rows.append((no, tipo, etapa, responsavel, ini_str, fim_str, entreg))

    # Identifica a ultima Fase root (No. sem ponto) para marcar como close
    last_fase_root_idx = None
    for idx, (no, tipo, *_rest) in enumerate(raw_rows):
        if tipo == "Fase" and "." not in no:
            last_fase_root_idx = idx

    for idx, (no, tipo, etapa, responsavel, ini_str, fim_str, entreg) in enumerate(raw_rows):
        if tipo == "Fase":
            stats["fases"] += 1
            close_cls = " close" if idx == last_fase_root_idx else ""
            rows_html_parts.append(
                f'      <tr class="phase-row{close_cls}">'
                f'<td class="wbs-code">{html_escape(no)}</td>'
                f'<td>{html_escape(tipo)}</td>'
                f'<td><strong>{html_escape(str(etapa))}</strong></td>'
                f'<td>{html_escape(str(responsavel))}</td>'
                f'<td>{html_escape(ini_str)}</td>'
                f'<td>{html_escape(fim_str)}</td>'
                f'<td class="entreg">{html_escape(str(entreg))}</td>'
                f'</tr>'
            )
        elif tipo in ("Ação", "Acao"):
            stats["acoes"] += 1
            rows_html_parts.append(
                f'      <tr class="acao-row">'
                f'<td class="wbs-code">{html_escape(no)}</td>'
                f'<td>{html_escape(tipo)}</td>'
                f'<td><strong>{html_escape(str(etapa))}</strong></td>'
                f'<td>{html_escape(str(responsavel))}</td>'
                f'<td>{html_escape(ini_str)}</td>'
                f'<td>{html_escape(fim_str)}</td>'
                f'<td class="entreg">{html_escape(str(entreg))}</td>'
                f'</tr>'
            )
        else:
            stats["etapas"] += 1
            rows_html_parts.append(
                f'      <tr class="etapa-row">'
                f'<td class="wbs-code sub">{html_escape(no)}</td>'
                f'<td>{html_escape(tipo)}</td>'
                f'<td>{html_escape(str(etapa))}</td>'
                f'<td>{html_escape(str(responsavel))}</td>'
                f'<td>{html_escape(ini_str)}</td>'
                f'<td>{html_escape(fim_str)}</td>'
                f'<td class="entreg-etapa">{html_escape(str(entreg))}</td>'
                f'</tr>'
            )

    period_start = data.get("period_start")
    period_end = data.get("period_end")
    dias = data.get("dias") or days_between(period_start, period_end)

    subs = shared_subs(
        data, logo_b64, "08", "Cronograma Detalhado",
        prev=("riscos.html", "Riscos"),
        next_=("calendario.html", "Próximo: Calendário"),
    )
    subs.update({
        "{{stat_fases}}": str(stats["fases"]),
        "{{stat_acoes}}": str(stats["acoes"]),
        "{{stat_etapas}}": str(stats["etapas"]),
        "{{stat_dias}}": str(dias),
        "{{cronograma_rows_html}}": "\n".join(rows_html_parts),
    })
    html = substitute(template, subs)
    output_path = output_dir / "artefatos" / "cronograma.html"
    write_html(output_path, html, warnings, "cronograma.html")
    return output_path


# ---------------------------------------------------------------------------
# Renderer 09 — calendario.html (consome events JSON gerado por derive_calendar_events.py)
# ---------------------------------------------------------------------------

def render_calendario(data: dict, logo_b64: str, output_dir: Path,
                      warnings: list[str]) -> Path:
    template = load_template("artefatos/calendario.tmpl.html")

    events = data.get("events") or []
    period_start = parse_date(data.get("period_start"))
    period_end = parse_date(data.get("period_end"))

    # Build MONTHS array
    months = []
    if period_start and period_end:
        cur = dt.datetime(period_start.year, period_start.month, 1)
        end_anchor = dt.datetime(period_end.year, period_end.month, 1)
        while cur <= end_anchor:
            months.append({
                "year": cur.year,
                "month": cur.month - 1,  # JS 0-indexed
                "name": MESES_BR_FULL[cur.month],
                "phases": [],
            })
            if cur.month == 12:
                cur = dt.datetime(cur.year + 1, 1, 1)
            else:
                cur = dt.datetime(cur.year, cur.month + 1, 1)

    # Stats cards
    counts = {}
    for ev in events:
        t = ev.get("type", "")
        counts[t] = counts.get(t, 0) + 1
    stats_cards_html = ""
    for t, label in [
        ("discovery", "Sessões Discovery"),
        ("validacao", "Validações"),
        ("checkin", "Check-ins"),
        ("status-report", "Status Reports"),
        ("handoff", "Handoffs"),
        ("marco", "Marcos"),
    ]:
        n = counts.get(t, 0)
        if n > 0:
            stats_cards_html += (
                f'    <div class="stat-card"><div class="val">{n}</div>'
                f'<div class="label">{html_escape(label)}</div></div>\n'
            )

    subs = shared_subs(
        data, logo_b64, "09", "Calendário de Reuniões & Rituais",
        prev=("cronograma.html", "Cronograma"),
        next_=("../plano-projeto.html", "Plano do Projeto"),
    )
    subs.update({
        "{{stats_cards_html}}": stats_cards_html.rstrip(),
        "{{events_json}}": json.dumps(events, ensure_ascii=False),
        "{{months_json}}": json.dumps(months, ensure_ascii=False),
    })
    html = substitute(template, subs)
    output_path = output_dir / "artefatos" / "calendario.html"
    write_html(output_path, html, warnings, "calendario.html")
    return output_path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    p = argparse.ArgumentParser(description="Renderiza os 10 HTMLs do plano de projeto.")
    p.add_argument("--input", help="JSON file com dados do plano (default: stdin)")
    p.add_argument("--output", required=True, help="Diretorio de saida (1-planning/)")
    p.add_argument("--xlsx", required=True, help="Path do Cronograma.xlsx (BASELINE) — necessario para artefato 08")
    args = p.parse_args()

    if args.input:
        try:
            data = json.loads(Path(args.input).read_text(encoding="utf-8"))
        except (FileNotFoundError, json.JSONDecodeError) as e:
            print(f"ERRO: nao consegui ler --input: {e}", file=sys.stderr)
            return 1
    else:
        try:
            data = json.loads(sys.stdin.read())
        except json.JSONDecodeError as e:
            print(f"ERRO: stdin nao e JSON valido: {e}", file=sys.stderr)
            return 1

    output_dir = Path(args.output)
    xlsx_path = Path(args.xlsx)
    warnings: list[str] = []
    files_written: list[str] = []

    try:
        logo_b64 = load_logo_b64("offwhite")
    except (FileNotFoundError, ValueError) as e:
        print(f"ERRO carregando logo: {e}", file=sys.stderr)
        return 1

    renderers = [
        ("00 landing",          lambda: render_landing(data, logo_b64, output_dir, warnings)),
        ("01 contexto-escopo",  lambda: render_contexto_escopo(data, logo_b64, output_dir, warnings)),
        ("02 eap",              lambda: render_eap(data, logo_b64, output_dir, warnings)),
        ("03 roadmap-marcos",   lambda: render_roadmap_marcos(data, logo_b64, output_dir, warnings)),
        ("04 okrs",             lambda: render_okrs(data, logo_b64, output_dir, warnings)),
        ("05 recursos",         lambda: render_recursos(data, logo_b64, output_dir, warnings)),
        ("06 plano-comunicacao", lambda: render_plano_comunicacao(data, logo_b64, output_dir, warnings)),
        ("07 riscos",           lambda: render_riscos(data, logo_b64, output_dir, warnings)),
        ("08 cronograma",       lambda: render_cronograma(data, logo_b64, xlsx_path, output_dir, warnings)),
        ("09 calendario",       lambda: render_calendario(data, logo_b64, output_dir, warnings)),
    ]

    for label, fn in renderers:
        try:
            path = fn()
            files_written.append(str(path))
        except Exception as e:
            print(f"ERRO em {label}: {e}", file=sys.stderr)
            return 1

    payload = {
        "ok": True,
        "files_written": files_written,
        "warnings": warnings,
        "n_files": len(files_written),
    }
    json.dump(payload, sys.stdout, ensure_ascii=False, indent=2)
    sys.stdout.write("\n")
    return 0


if __name__ == "__main__":
    sys.exit(main())
