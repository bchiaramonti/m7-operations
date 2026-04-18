#!/usr/bin/env python3
"""
generate_xlsx.py — Gera Cronograma.xlsx (BASELINE) do plano de projeto.

Produz `<project-dir>/1-planning/Cronograma.xlsx` reproduzindo o schema do
projeto-modelo H1-02 com tokens M7-2026:
    - Sheet único `Cronograma Detalhado`
    - Linhas 1-3: titulo + subtitulo do projeto + linha em branco
    - Linha 4: header (10 colunas B-K) com fundo caqui + texto off-white + bold
    - Linha 5+: dados, formatação por Tipo:
        Fase  -> fundo caqui #424135, texto off-white, bold
        Ação  -> fundo cinza-200 #d9d9d9, texto preto
        Etapa -> fundo branco, texto cinza-700, fonte 10pt
    - Datas (cols F, G, H, I): formato dd/mm/aaaa
    - Data validation em col C (Tipo) e col J (Status)
    - Freeze panes em A5
    - Auto-filter na row 4
    - Status sempre `not_started` na BASELINE; Início Real / Fim Real vazios
    - **Sem coluna L `ClickUp ID`** (adicionada por managing-action-plan no 1º run)

Input: JSON via stdin OU arquivo (--input) com schema:
    {
      "project_name": "Playbook de Processos",
      "project_code": "H1-02",
      "project_subtitle": "90 dias | 27/mar a 18/jul 2026 | ...",
      "rows": [
        {"no": "1", "tipo": "Fase", "etapa": "FASE 1...", "responsavel": "Bruno",
         "inicio_plan": "2026-03-27", "fim_plan": "2026-04-14", "entregavel": ""},
        ...
      ]
    }

Saida: arquivo xlsx em --output (path) + JSON em stdout com {ok, path, rows_written}.

Uso:
    cat plan-data.json | python3 generate_xlsx.py --output 1-planning/Cronograma.xlsx
    python3 generate_xlsx.py --input data.json --output 1-planning/Cronograma.xlsx
"""

import argparse
import datetime as dt
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("ERRO: openpyxl nao instalado. Execute: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)


# ---------------------------------------------------------------------------
# Tokens M7-2026 (replicados em codigo para evitar dependencia externa)
# ---------------------------------------------------------------------------

VERDE_CAQUI = "FF424135"     # M7 verde caqui (sem #)
OFF_WHITE = "FFFFFDEF"
CINZA_200 = "FFD9D9C8"
CINZA_100 = "FFF0F0E4"
CINZA_700 = "FF3D3D3D"
LIME = "FFEEF77C"

SHEET_NAME = "Cronograma Detalhado"
HEADER_ROW = 4
FIRST_DATA_ROW = 5

# Map nome canonico -> (column letter, header label, width)
COLUMNS = [
    ("no",          "B", "No.",               7),
    ("tipo",        "C", "Tipo",              16),
    ("etapa",       "D", "Etapa",             70),
    ("responsavel", "E", "Responsável",       18),
    ("inicio_plan", "F", "Início Planejado",  16),
    ("fim_plan",    "G", "Fim Planejado",     16),
    ("inicio_real", "H", "Início Real",       14),
    ("fim_real",    "I", "Fim Real",          14),
    ("status",      "J", "Status",            14),
    ("entregavel",  "K", "Entregável",        70),
]

TIPOS_VALIDOS = ["Fase", "Ação", "Etapas da Ação"]
STATUS_VALIDOS = ["not_started", "in_progress", "blocked", "done"]

ISO_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def parse_date(value):
    """Aceita datetime, ISO string ou retorna None."""
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        return value
    if isinstance(value, dt.date):
        return dt.datetime(value.year, value.month, value.day)
    if isinstance(value, str):
        if ISO_DATE_RE.match(value):
            return dt.datetime.strptime(value, "%Y-%m-%d")
    return None


def is_fase_row(row):
    return str(row.get("tipo", "")).strip() in ("Fase",)


def is_acao_row(row):
    return str(row.get("tipo", "")).strip() in ("Ação", "Acao")


def is_etapa_row(row):
    return str(row.get("tipo", "")).strip() in ("Etapas da Ação", "Etapas da Acao")


def style_header_cell(cell):
    cell.font = Font(name="Calibri", size=11, bold=True, color=OFF_WHITE)
    cell.fill = PatternFill("solid", fgColor=VERDE_CAQUI)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_fase_row(cell):
    cell.font = Font(name="Calibri", size=11, bold=True, color=OFF_WHITE)
    cell.fill = PatternFill("solid", fgColor=VERDE_CAQUI)
    cell.alignment = Alignment(vertical="center", wrap_text=True)


def style_acao_row(cell):
    cell.font = Font(name="Calibri", size=11, bold=False, color="FF000000")
    cell.fill = PatternFill("solid", fgColor="FFD9D9D9")
    cell.alignment = Alignment(vertical="center", wrap_text=True)


def style_etapa_row(cell):
    cell.font = Font(name="Calibri", size=10, bold=False, color=CINZA_700)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    # sem fill explicito (branco padrao)


def style_title_cell(cell, size=16):
    cell.font = Font(name="Calibri", size=size, bold=True, color=VERDE_CAQUI)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_subtitle_cell(cell):
    cell.font = Font(name="Calibri", size=11, italic=True, color="FF666666")
    cell.alignment = Alignment(horizontal="left", vertical="center")


def col_letter(canon_name):
    for name, letter, _, _ in COLUMNS:
        if name == canon_name:
            return letter
    raise ValueError(f"Coluna canonica desconhecida: {canon_name}")


def write_xlsx(payload, output_path):
    project_name = payload.get("project_name", "(sem nome)")
    project_code = payload.get("project_code", "")
    subtitle = payload.get("project_subtitle", "")
    rows = payload.get("rows", [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # --- Title rows (1-3) ---
    title_text = f"Cronograma Detalhado — {project_code} {project_name}".strip()
    ws.cell(1, 3, title_text)
    style_title_cell(ws.cell(1, 3), size=16)
    if subtitle:
        ws.cell(2, 3, subtitle)
        style_subtitle_cell(ws.cell(2, 3))
    # row 3 fica vazia (respiro visual)

    # --- Header (row 4) ---
    for canon, letter, label, _width in COLUMNS:
        cell = ws.cell(HEADER_ROW, openpyxl.utils.column_index_from_string(letter), label)
        style_header_cell(cell)
    ws.row_dimensions[HEADER_ROW].height = 30

    # --- Column widths ---
    ws.column_dimensions["A"].width = 4  # margem
    for canon, letter, _label, width in COLUMNS:
        ws.column_dimensions[letter].width = width

    # --- Data rows (row 5+) ---
    rows_written = 0
    for i, row in enumerate(rows):
        r = FIRST_DATA_ROW + i

        # Pre-determinar styler
        if is_fase_row(row):
            styler = style_fase_row
            ws.row_dimensions[r].height = 22
        elif is_acao_row(row):
            styler = style_acao_row
            ws.row_dimensions[r].height = 22
        else:
            styler = style_etapa_row
            ws.row_dimensions[r].height = 18

        for canon, letter, _label, _width in COLUMNS:
            col_idx = openpyxl.utils.column_index_from_string(letter)
            value = row.get(canon, "")
            cell = ws.cell(r, col_idx)

            # Valor
            if canon in ("inicio_plan", "fim_plan", "inicio_real", "fim_real"):
                dt_val = parse_date(value)
                if dt_val:
                    cell.value = dt_val
                    cell.number_format = "dd/mm/yyyy"
            elif canon == "status":
                # BASELINE: Status sempre not_started
                cell.value = row.get("status") or "not_started"
            else:
                if value is not None and value != "":
                    cell.value = value

            # Aplicar style
            styler(cell)
            # Bold no titulo da etapa para acoes (espelhando H1-02)
            if canon == "etapa" and is_acao_row(row):
                cell.font = Font(name="Calibri", size=11, bold=True, color="FF000000")

        rows_written += 1

    # --- Freeze panes ---
    ws.freeze_panes = ws.cell(FIRST_DATA_ROW, openpyxl.utils.column_index_from_string("B"))

    # --- Auto-filter ---
    last_data_row = FIRST_DATA_ROW + len(rows) - 1 if rows else FIRST_DATA_ROW
    last_col_letter = COLUMNS[-1][1]
    ws.auto_filter.ref = f"B{HEADER_ROW}:{last_col_letter}{max(last_data_row, HEADER_ROW)}"

    # --- Data validation ---
    if rows:
        dv_tipo = DataValidation(
            type="list",
            formula1=f'"{",".join(TIPOS_VALIDOS)}"',
            allow_blank=False,
            showErrorMessage=True,
            errorTitle="Tipo invalido",
            error="Use: Fase | Ação | Etapas da Ação",
        )
        dv_status = DataValidation(
            type="list",
            formula1=f'"{",".join(STATUS_VALIDOS)}"',
            allow_blank=False,
            showErrorMessage=True,
            errorTitle="Status invalido",
            error="Use: not_started | in_progress | blocked | done",
        )
        ws.add_data_validation(dv_tipo)
        ws.add_data_validation(dv_status)
        dv_tipo.add(f"C{FIRST_DATA_ROW}:C{last_data_row}")
        dv_status.add(f"J{FIRST_DATA_ROW}:J{last_data_row}")

    # --- Save ---
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return rows_written


def validate_payload(payload):
    """Valida integridade basica antes de gerar."""
    errors = []
    rows = payload.get("rows", [])
    if not rows:
        errors.append("payload sem rows (precisa pelo menos 1 linha de dados)")

    seen_no = set()
    for i, row in enumerate(rows, start=1):
        no = str(row.get("no", "") or "").strip()
        if not no:
            errors.append(f"row {i}: campo 'no' obrigatorio")
        elif no in seen_no:
            errors.append(f"row {i}: No. duplicado '{no}'")
        else:
            seen_no.add(no)

        if not row.get("tipo"):
            errors.append(f"row {i} (No. {no}): campo 'tipo' obrigatorio")
        elif row["tipo"] not in TIPOS_VALIDOS and row["tipo"] not in ("Acao", "Etapas da Acao"):
            errors.append(f"row {i} (No. {no}): tipo '{row['tipo']}' invalido")

        if not row.get("etapa"):
            errors.append(f"row {i} (No. {no}): campo 'etapa' obrigatorio")

        # Datas planejadas obrigatorias
        for date_field in ("inicio_plan", "fim_plan"):
            v = row.get(date_field)
            if not v:
                errors.append(f"row {i} (No. {no}): campo '{date_field}' obrigatorio")

        # inicio <= fim
        ini = parse_date(row.get("inicio_plan"))
        fim = parse_date(row.get("fim_plan"))
        if ini and fim and ini > fim:
            errors.append(f"row {i} (No. {no}): inicio_plan > fim_plan")

    return errors


def main() -> int:
    p = argparse.ArgumentParser(description="Gera Cronograma.xlsx BASELINE.")
    p.add_argument("--input", help="JSON file com schema do plano (default: stdin)")
    p.add_argument("--output", required=True, help="Path do xlsx de saida")
    p.add_argument("--validate-only", action="store_true",
                   help="So valida payload sem gerar arquivo")
    args = p.parse_args()

    # Carrega payload
    if args.input:
        try:
            payload = json.loads(Path(args.input).read_text(encoding="utf-8"))
        except (FileNotFoundError, json.JSONDecodeError) as e:
            print(f"ERRO: nao consegui ler --input: {e}", file=sys.stderr)
            return 1
    else:
        try:
            payload = json.loads(sys.stdin.read())
        except json.JSONDecodeError as e:
            print(f"ERRO: stdin nao e JSON valido: {e}", file=sys.stderr)
            return 1

    # Valida
    errors = validate_payload(payload)
    if errors:
        print(f"ERRO: payload invalido ({len(errors)} erros):", file=sys.stderr)
        for e in errors[:10]:
            print(f"  - {e}", file=sys.stderr)
        if len(errors) > 10:
            print(f"  ... e mais {len(errors) - 10} erros", file=sys.stderr)
        return 1

    if args.validate_only:
        print(json.dumps({"ok": True, "rows": len(payload.get("rows", []))}))
        return 0

    # Gera
    try:
        rows_written = write_xlsx(payload, args.output)
    except Exception as e:
        print(f"ERRO ao gerar xlsx: {e}", file=sys.stderr)
        return 1

    print(json.dumps({
        "ok": True,
        "path": args.output,
        "rows_written": rows_written,
        "sheet": SHEET_NAME,
    }))
    return 0


if __name__ == "__main__":
    sys.exit(main())
