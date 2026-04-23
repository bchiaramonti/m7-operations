#!/usr/bin/env python3
"""
collect_data.py — Deterministic data collection for generating-status-materials.

Reads:
  - <proj>/CLAUDE.md
  - <proj>/BRIEFING.md (if present)
  - <proj>/1-planning/plano-projeto.html + artefatos/*.html
  - <proj>/4-status-report/Cronograma.xlsx (LIVE)
  - <proj>/4-status-report/changelog.md
  - <proj>/4-status-report/.sync-state.json (optional)

Emits a canonical JSON dict on stdout consumed by build_opr.py and build_pptx.py.

Pure Python: no MCP calls, no LLM inference. All narrative synthesis is
deterministic (rules in references/narrative-synthesis.md).
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any


CANONICAL_COLUMNS = {
    "No.": "no",
    "Tipo": "tipo",
    "Etapa": "etapa",
    "Responsável": "responsavel",
    "Responsavel": "responsavel",
    "Início Planejado": "inicio_plan",
    "Inicio Planejado": "inicio_plan",
    "Fim Planejado": "fim_plan",
    "Início Real": "inicio_real",
    "Inicio Real": "inicio_real",
    "Fim Real": "fim_real",
    "Status": "status",
    "Entregável": "entregavel",
    "Entregavel": "entregavel",
    "ClickUp ID": "clickup_id",
}
REQUIRED_COLUMNS = {"no", "tipo", "etapa", "status", "fim_plan"}

DATE_FORMATS = ("%Y-%m-%d", "%d/%m/%Y", "%d/%m", "%d/%b/%Y", "%d/%b")

MONTH_PT = {
    1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
    5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
    9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro",
}

# Portuguese abbreviated month names → English (locale "C" uses English %b).
# Used to pre-translate dates like "14/abr" before parsing.
MONTH_PT_ABBREV_TO_EN = {
    "jan": "jan", "fev": "feb", "mar": "mar", "abr": "apr",
    "mai": "may", "jun": "jun", "jul": "jul", "ago": "aug",
    "set": "sep", "out": "oct", "nov": "nov", "dez": "dec",
}


@dataclass
class CollectResult:
    report_date: str
    project: dict = field(default_factory=dict)
    status: dict = field(default_factory=dict)
    highlights: list = field(default_factory=list)
    next_steps: list = field(default_factory=list)
    progress_concluidas: list = field(default_factory=list)
    progress_proximas: list = field(default_factory=list)
    attentions: list = field(default_factory=list)
    milestones: list = field(default_factory=list)
    macro_milestones: list = field(default_factory=list)
    sprints: list = field(default_factory=list)
    fronts: list = field(default_factory=list)
    sprint_actions: list = field(default_factory=list)
    risks: list = field(default_factory=list)
    changelog_entries: list = field(default_factory=list)
    roadmap_overlays: dict = field(default_factory=dict)
    roadmap_structure: dict = field(default_factory=dict)
    warnings: list = field(default_factory=list)

    def to_dict(self) -> dict:
        return {
            "report_date": self.report_date,
            "project": self.project,
            "status": self.status,
            "highlights": self.highlights,
            "next_steps": self.next_steps,
            "progress_concluidas": self.progress_concluidas,
            "progress_proximas": self.progress_proximas,
            "attentions": self.attentions,
            "milestones": self.milestones,
            "macro_milestones": self.macro_milestones,
            "sprints": self.sprints,
            "fronts": self.fronts,
            "sprint_actions": self.sprint_actions,
            "risks": self.risks,
            "changelog_entries": self.changelog_entries,
            "roadmap_overlays": self.roadmap_overlays,
            "roadmap_structure": self.roadmap_structure,
            "warnings": self.warnings,
        }


def normalize_date(value, default_year: int) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        s = value.strip()
        if s in ("—", "-", ""):
            return None
        # Pre-translate Portuguese month abbreviations to English to match %b in C locale
        s_translated = s
        for pt, en in MONTH_PT_ABBREV_TO_EN.items():
            # Case-insensitive replace of 3-letter PT abbrev when bordered by / or end
            s_translated = re.sub(
                rf"(?<=/){pt}(?=/|$)",
                en,
                s_translated,
                flags=re.IGNORECASE,
            )
        for candidate in (s, s_translated):
            for fmt in DATE_FORMATS:
                try:
                    dt = datetime.strptime(candidate, fmt)
                    if dt.year == 1900:
                        dt = dt.replace(year=default_year)
                    return dt
                except ValueError:
                    continue
    return None


def parent_no(no: str) -> str | None:
    parts = str(no).split(".")
    if len(parts) <= 1:
        return None
    return ".".join(parts[:-1])


# ---------- Readers ----------

def read_cronograma_live(path: Path, default_year: int, warnings: list) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl ausente. Instale: pip install openpyxl")

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 5:
        raise ValueError("Cronograma.xlsx com menos de 5 linhas — schema esperado tem header em R4.")

    header_row = rows[3]
    col_map: dict[int, str] = {}
    for idx, val in enumerate(header_row):
        if isinstance(val, str) and val.strip() in CANONICAL_COLUMNS:
            col_map[idx] = CANONICAL_COLUMNS[val.strip()]

    present = set(col_map.values())
    missing_required = REQUIRED_COLUMNS - present
    if missing_required:
        raise ValueError(
            f"Cronograma.xlsx faltando colunas obrigatórias: {sorted(missing_required)}"
        )

    entries = []
    for row_num, row in enumerate(rows[4:], start=5):
        if not row or all(v is None for v in row):
            continue
        entry: dict[str, Any] = {"_row": row_num}
        for idx, canonical in col_map.items():
            if idx < len(row):
                entry[canonical] = row[idx]
        if not entry.get("no"):
            continue
        # Normalize dates
        for date_field in ("inicio_plan", "fim_plan", "inicio_real", "fim_real"):
            if date_field in entry:
                entry[date_field] = normalize_date(entry[date_field], default_year)
        # Normalize status
        status = entry.get("status")
        if status and status not in ("not_started", "in_progress", "blocked", "done"):
            warnings.append(f"Cronograma linha {row_num}: status '{status}' fora do enum — tratando como not_started")
            entry["status"] = "not_started"
        entries.append(entry)
    return entries


def read_changelog(path: Path) -> list[dict]:
    if not path.exists():
        return []
    text = path.read_text(encoding="utf-8")
    entries = []
    # Pattern: ## YYYY-MM-DD HH:MM — op
    header_pat = re.compile(r"^##\s+(\d{4}-\d{2}-\d{2})(?:[T ](\d{2}:\d{2}))?\s*[—-]\s*(\w+)", re.MULTILINE)
    matches = list(header_pat.finditer(text))
    for i, m in enumerate(matches):
        date_str, time_str, op = m.group(1), m.group(2) or "00:00", m.group(3)
        ts = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        body = text[m.end():end].strip()
        entry = {"timestamp": ts.isoformat(), "op": op, "text": body, "raw": body}
        # Try to extract common fields
        for field_name in ("No.", "Campo", "Antes", "Depois", "Por"):
            pat = re.compile(rf"\*\*{re.escape(field_name)}:\*\*\s*(.+)")
            mm = pat.search(body)
            if mm:
                key = {"No.": "no", "Campo": "field", "Antes": "old", "Depois": "new", "Por": "by"}[field_name]
                entry[key] = mm.group(1).strip()
        entries.append(entry)
    return entries


def read_sync_state(path: Path, report_date: datetime, warnings: list) -> dict:
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as e:
        warnings.append(f".sync-state.json inválido: {e}")
        return {}
    if data.get("sync_pending"):
        warnings.append("sync_pending=true — execute `managing-action-plan sync` antes do reporte para dados frescos")
    last_sync = data.get("last_sync")
    if last_sync:
        try:
            last_dt = datetime.fromisoformat(last_sync)
            age = report_date - last_dt
            if age > timedelta(hours=48):
                hours = int(age.total_seconds() / 3600)
                warnings.append(f"Sync stale: última sync há {hours}h ({last_sync})")
        except ValueError:
            pass
    return data


def read_html_text(path: Path, warnings: list, label: str) -> dict:
    """Loosely parses an HTML and returns {title, text, metadata}."""
    if not path.exists():
        warnings.append(f"{label}: arquivo não encontrado em {path}")
        return {}
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        raise RuntimeError("beautifulsoup4 ausente. Instale: pip install beautifulsoup4")
    soup = BeautifulSoup(path.read_text(encoding="utf-8"), "html.parser")
    title_tag = soup.find("h1") or soup.find("title")
    return {
        "title": title_tag.get_text(strip=True) if title_tag else None,
        "soup": soup,
    }


def parse_risks(path: Path, warnings: list) -> list[dict]:
    """Parses .risk-item cards from the riscos.html artifact produced by building-project-plan.

    Schema extracted per card:
      - .risk-id (text = code like "R01"/"O01"; class in {crit, high, med, low} = severity)
      - .risk-content > h4 = title
      - .risk-content > .tags > .tag-prob = "Prob: Alta/Média/Baixa"
      - .risk-content > .tags > .tag-imp  = "Imp: Crítico/Alto/Médio/Baixo"
      - .risk-content > p                 = description
      - .risk-content > .mitigation > span = contramedida
    """
    if not path.exists():
        warnings.append("riscos.html não encontrado — seção Riscos vazia")
        return []
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        return []
    soup = BeautifulSoup(path.read_text(encoding="utf-8"), "html.parser")
    risks = []

    sev_class_map = {"crit": "critico", "high": "alto", "med": "medio", "low": "baixo"}

    for idx, card in enumerate(soup.select(".risk-item"), 1):
        id_el = card.select_one(".risk-id")
        code = id_el.get_text(strip=True) if id_el else f"R{idx:02d}"

        # Severity comes from the class of .risk-id (crit/high/med/low).
        # This is the authoritative visual severity from the plan.
        sev_raw = "low"
        if id_el:
            for c in id_el.get("class", []):
                if c in sev_class_map:
                    sev_raw = c
                    break
        severity_label = sev_class_map[sev_raw]

        title_el = card.select_one(".risk-content h4") or card.select_one("h4")
        title = title_el.get_text(strip=True) if title_el else ""

        prob_el = card.select_one(".tag-prob")
        prob_text = prob_el.get_text(strip=True) if prob_el else ""
        prob_value = _extract_after_colon(prob_text)

        imp_el = card.select_one(".tag-imp")
        imp_text = imp_el.get_text(strip=True) if imp_el else ""
        imp_value = _extract_after_colon(imp_text)

        desc_el = card.select_one(".risk-content > p")
        description = desc_el.get_text(" ", strip=True) if desc_el else ""

        mitig_el = card.select_one(".mitigation span") or card.select_one(".mitigation")
        mitigation = ""
        if mitig_el:
            raw = mitig_el.get_text(" ", strip=True)
            # Remove leading "Ação de Mitigação" header if present
            mitigation = re.sub(r"^Ação de Mitigação\s*", "", raw).strip()

        is_upside = code.upper().startswith("O")

        # Incurred = risco que saiu de "monitoramento" para "incidente ativo".
        # Opt-in explícito via `data-incurred="true"` no .risk-item do HTML de planejamento
        # (building-project-plan pode expor isso no futuro). Hoje o default é False — o OPR
        # exibe placeholder minimalista quando não há risco incorrido.
        incurred_attr = (card.get("data-incurred") or "").strip().lower()
        is_incurred = incurred_attr in ("true", "1", "yes", "sim")

        risks.append({
            "code": code,
            "title": title,
            "probability": _normalize_level(prob_value),
            "impact": _normalize_impact(imp_value),
            "severity": severity_label,
            "severity_class": sev_raw,
            "description": description,
            "mitigation": mitigation,
            "is_upside": is_upside,
            "is_incurred": is_incurred,
        })

    if not risks:
        warnings.append("riscos.html: nenhum .risk-item encontrado — verifique se o arquivo foi gerado por building-project-plan")
    return risks


def _extract_after_colon(text: str) -> str:
    """Parses labels like 'Prob: Alta' or 'Imp: Crítico' -> 'alta' / 'crítico'."""
    if ":" in text:
        return text.split(":", 1)[1].strip().lower()
    return text.strip().lower()


def _normalize_level(s: str) -> str:
    s = s.strip().lower()
    if "alt" in s:
        return "alta"
    if "med" in s or "méd" in s:
        return "media"
    return "baixa"


def _normalize_impact(s: str) -> str:
    s = s.strip().lower()
    if "crít" in s or "crit" in s:
        return "critico"
    if "alt" in s:
        return "alto"
    if "med" in s or "méd" in s:
        return "medio"
    return "baixo"


def parse_plano_projeto(path: Path, warnings: list) -> dict:
    data = read_html_text(path, warnings, "plano-projeto.html")
    if not data:
        return {}
    soup = data["soup"]
    result = {"name": data.get("title")}
    # Common patterns: .hero-meta data-key="pm", .hero-meta__pm, etc.
    for key in ("pm", "lider", "líder", "sponsor", "duracao", "duração", "goal", "objetivo", "estrela"):
        el = soup.select_one(f"[data-key='{key}']") or soup.select_one(f".hero-meta__{key}")
        if el:
            result_key = {"lider": "pm", "líder": "pm", "objetivo": "goal"}.get(key, key)
            result[result_key] = el.get_text(strip=True)
    return result


def parse_milestones(path: Path, warnings: list, default_year: int, report_date: datetime) -> list[dict]:
    """Parses project milestones from roadmap-marcos.html.

    Looks for the canonical structure produced by building-project-plan:
      .lane.milestones > .track > .tick.{top|bottom}.{gate|regular}
        ├ .lbl  ("M0 KICKOFF")
        ├ .date ("14/abr")
        └ .desc ("TAP + OKRs aprovados")

    Status is inferred deterministically by date:
      - date <= report_date  → "done"   (treat as passed)
      - date == report_date ± 3 days → "in_progress"
      - date > report_date   → "not_started"

    `major` flag is True for .gate ticks (critical gates), False for .regular (informational).
    """
    if not path.exists():
        warnings.append("roadmap-marcos.html não encontrado — lista de marcos vazia")
        return []
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        return []
    soup = BeautifulSoup(path.read_text(encoding="utf-8"), "html.parser")

    milestones = []
    # Primary: ticks inside the milestones lane
    for tick in soup.select(".lane.milestones .tick"):
        lbl_el = tick.select_one(".lbl")
        date_el = tick.select_one(".date")
        desc_el = tick.select_one(".desc")
        if not lbl_el or not date_el:
            continue
        label = lbl_el.get_text(" ", strip=True)
        date_str = date_el.get_text(strip=True)
        dt = normalize_date(date_str, default_year)

        # Classes: tick top|bottom gate|regular
        classes = tick.get("class", [])
        is_gate = "gate" in classes

        # Inline style: "left:15.79%;" — authoritative anchor for date→pct calibration
        left_pct = None
        style = tick.get("style", "")
        m_left = re.search(r"left\s*:\s*([\d.]+)\s*%", style)
        if m_left:
            try:
                left_pct = float(m_left.group(1))
            except ValueError:
                left_pct = None

        if dt is None:
            status = "not_started"
        else:
            delta_days = (report_date - dt).days
            if delta_days > 3:
                status = "done"
            elif delta_days >= -3:
                status = "in_progress"
            else:
                status = "not_started"

        milestones.append({
            "code": label.split()[0] if label else "",
            "label": label,
            "name": label,
            "date": dt.strftime("%Y-%m-%d") if dt else None,
            "date_short": date_str,
            "description": desc_el.get_text(" ", strip=True) if desc_el else "",
            "status": status,
            "is_critical": is_gate,
            "major": is_gate,
            "left_pct": left_pct,
        })

    # Fallback: generic .milestone lookup (legacy HTML structure)
    if not milestones:
        for item in soup.select(".milestone, [data-milestone], li.milestone"):
            name_el = item.select_one(".name, .title, h4")
            date_el = item.select_one(".date, [data-key='date'], time")
            if not name_el:
                continue
            date_val = None
            if date_el:
                date_val = date_el.get("datetime") or date_el.get_text(strip=True)
            dt = normalize_date(date_val, default_year) if date_val else None
            status = "not_started"
            if dt:
                delta = (report_date - dt).days
                status = "done" if delta > 3 else ("in_progress" if delta >= -3 else "not_started")
            milestones.append({
                "code": "",
                "label": name_el.get_text(strip=True),
                "name": name_el.get_text(strip=True),
                "date": dt.strftime("%Y-%m-%d") if dt else None,
                "date_short": date_val if isinstance(date_val, str) else None,
                "description": "",
                "status": status,
                "is_critical": "critical" in (item.get("class") or []),
                "major": "critical" in (item.get("class") or []),
            })

    return milestones


def _collect_bar_titles(roadmap_html_path: Path) -> list[str]:
    """Extracts unique `.bar .title` texts from the swim-lane roadmap HTML.
    Returns the list preserving document order (useful for debugging)."""
    if not roadmap_html_path.exists():
        return []
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        return []
    soup = BeautifulSoup(roadmap_html_path.read_text(encoding="utf-8"), "html.parser")
    titles = []
    seen = set()
    for title_el in soup.select(".bar .title"):
        txt = title_el.get_text(" ", strip=True)
        if txt and txt not in seen:
            seen.add(txt)
            titles.append(txt)
    return titles


def _normalize_match(s: str) -> str:
    """Normalization for fuzzy title↔etapa matching: lowercase + strip accents + collapse whitespace."""
    import unicodedata
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s.lower().strip())


def aggregate_bar_status(bar_title: str, entries: list[dict], report_date: datetime) -> str:
    """Aggregates execution status for a roadmap bar by matching its title against
    the `etapa` field of tasks in Cronograma.xlsx.

    Returns one of: "done", "active", "overdue", "future".

    Matching strategy (cascades from strict to lenient):
      1. Full normalized title as substring of etapa (e.g., "investment banking" in "…Investment Banking")
      2. If no matches and title has separators (·, •, —, –, +, &, /), split into tokens and
         try each token. Any token with ≥2 chars that matches contributes to the pool.

    This handles compound bar titles like "TAP · WBS · OKRs · Riscos" where individual
    tokens (TAP, WBS, OKRs, Riscos) each correspond to separate tasks in the xlsx.
    """
    bar_norm = _normalize_match(bar_title)
    if not bar_norm:
        return "future"

    acoes = [e for e in entries if str(e.get("tipo", "")).lower() in ("ação", "acao")]

    # Strategy 1: full title match
    matches = [e for e in acoes if bar_norm in _normalize_match(str(e.get("etapa", "")))]

    # Strategy 2: split on separators and try tokens (only if no full match).
    # Use min length 3 and a stopword filter to avoid false positives like "TE", "de", "da".
    _STOP = {"de", "da", "do", "das", "dos", "e", "a", "o", "as", "os", "na", "no", "em"}
    if not matches:
        tokens = [
            tok.strip() for tok in re.split(r"[·•\-–—+&/|]| e |,", bar_title)
            if tok.strip()
        ]
        seen_ids = set()
        for tok in tokens:
            tok_norm = _normalize_match(tok)
            if not tok_norm or len(tok_norm) < 3 or tok_norm in _STOP:
                continue
            for e in acoes:
                eid = id(e)
                if eid in seen_ids:
                    continue
                if tok_norm in _normalize_match(str(e.get("etapa", ""))):
                    matches.append(e)
                    seen_ids.add(eid)

    if not matches:
        return "future"

    done_count = sum(1 for e in matches if e.get("status") == "done")
    any_in_progress = any(e.get("status") in ("in_progress", "blocked") for e in matches)
    total = len(matches)

    latest_fim = None
    for e in matches:
        fp = e.get("fim_plan")
        if isinstance(fp, datetime) and (latest_fim is None or fp > latest_fim):
            latest_fim = fp

    if done_count == total:
        return "done"
    if any_in_progress or done_count > 0:
        return "active"
    if latest_fim is not None and latest_fim < report_date:
        return "overdue"
    return "future"


def infer_matrix_structure(entries: list[dict], bar_titles: list[str]) -> dict | None:
    """Attempts to infer the (processo, fase) matrix structure from the project's WBS.

    Heuristic: uses `bar_titles` (from roadmap-marcos.html .bar .title elements) as
    the canonical list of processos. For each bar_title, collects Ação-type tasks
    whose `etapa` contains the title. If all processos have the same count N of
    matched tasks AND the tasks can be split into N uniform "fase" patterns
    (prefixes that appear across all processos), returns the inferred structure.

    Returns None when the project doesn't follow a uniform processo × fase pattern
    (signals the caller to fall back to user-driven configuration).
    """
    if not bar_titles:
        return None
    acoes = [e for e in entries if str(e.get("tipo", "")).lower() in ("ação", "acao")]
    if not acoes:
        return None

    # For each bar_title, collect matching tasks (ordered by WBS code)
    processo_tasks: dict[str, list[dict]] = {}
    for title in bar_titles:
        t_norm = _normalize_match(title)
        if not t_norm:
            continue
        matches = [
            e for e in acoes
            if t_norm in _normalize_match(str(e.get("etapa", "")))
        ]
        # Sort by "No." (WBS code) if numeric-like
        matches.sort(key=lambda e: str(e.get("no", "")))
        if matches:
            processo_tasks[title] = matches

    if len(processo_tasks) < 2:
        return None  # need at least 2 processos to triangulate common fases

    # Keep only processos with the modal task count (most common count)
    from collections import Counter
    counts = Counter(len(v) for v in processo_tasks.values())
    modal_count, modal_freq = counts.most_common(1)[0]
    if modal_count < 2 or modal_freq < len(processo_tasks) * 0.8:
        # Too much variance in task counts → can't assume uniform fases
        return None

    uniform_processos = {p: ts for p, ts in processo_tasks.items() if len(ts) == modal_count}
    if len(uniform_processos) < 2:
        return None

    # Extract "fase" pattern from task etapa: remove processo label from etapa, keep prefix
    # E.g. "Mapa N2 — Jornada do cliente Crédito" - "Crédito" = "Mapa N2 — Jornada do cliente"
    def strip_processo(etapa: str, processo: str) -> str:
        e_norm = _normalize_match(etapa)
        p_norm = _normalize_match(processo)
        idx = e_norm.find(p_norm)
        if idx == -1:
            return etapa.strip()
        # Find the same position in original (case-insensitive find on original)
        lower = etapa.lower()
        p_lower = processo.lower()
        real_idx = lower.find(p_lower)
        if real_idx == -1:
            return etapa.strip()
        return etapa[:real_idx].strip().rstrip("—–-·:").strip()

    # Build fase_slots: for position i (0..N-1), collect the stripped prefixes across processos
    fase_slots: list[list[str]] = [[] for _ in range(modal_count)]
    for processo, tasks in uniform_processos.items():
        for i, task in enumerate(tasks):
            prefix = strip_processo(str(task.get("etapa", "")), processo)
            fase_slots[i].append(prefix)

    # For each slot, if all (or nearly all) prefixes are identical, use that as the fase pattern
    fases = []
    for i, prefixes in enumerate(fase_slots):
        if not prefixes:
            return None
        most_common, freq = Counter(prefixes).most_common(1)[0]
        if freq < len(prefixes) * 0.7:
            # Inconsistent prefixes in this slot → structure isn't uniform
            return None
        if not most_common:
            return None
        # Short label: extract the most descriptive "head" of the pattern.
        # Strategy: first non-stopword token, with two overrides:
        #   (1) Generic openers (Plano, Mapa, Fase) → try second token:
        #       - if it's a code (N1, N2, P1...), concat: "Mapa N2"
        #       - else use it alone: "Plano de implementação" → "Implementação"
        #   (2) Short codes in second slot that look like N2/P1 → concat with first.
        # Examples:
        #   "Mapa N2 — Jornada do cliente" → "Mapa N2"
        #   "DEIP + Tabela de Desconexões" → "DEIP"
        #   "Políticas e manuais aplicáveis" → "Políticas"
        #   "Playbook consolidado" → "Playbook"
        #   "Plano de implementação" → "Implementação"
        _STOP_LABEL = {"de", "da", "do", "das", "dos", "e", "em", "a", "o",
                       "para", "por", "com", "sem", "na", "no"}
        _GENERIC = {"plano", "mapa", "fase", "etapa"}
        tokens = [t for t in re.split(r"[\s—–·+]+", most_common) if t]
        head_tokens: list[str] = []
        for tok in tokens:
            if tok.lower() in _STOP_LABEL:
                continue
            head_tokens.append(tok)
            if len(head_tokens) >= 2:
                break
        if not head_tokens:
            short_label = most_common.split()[0] if most_common.split() else most_common
        else:
            t0 = head_tokens[0]
            t0_is_generic = t0.lower() in _GENERIC
            if len(head_tokens) == 1:
                short_label = t0
            else:
                t1 = head_tokens[1]
                t1_is_code = (len(t1) <= 3 and t1[0].isalpha()
                              and t1[1:].replace(".", "").isdigit())
                if t1_is_code:
                    short_label = f"{t0} {t1}"
                elif t0_is_generic:
                    # Generic opener + noun: use noun with Title-casing
                    short_label = t1[:1].upper() + t1[1:]
                else:
                    short_label = t0

        # Code = normalized uppercase version of label (accent-stripped)
        import unicodedata
        code_base = unicodedata.normalize("NFD", short_label)
        code_base = "".join(c for c in code_base if unicodedata.category(c) != "Mn")
        code = re.sub(r"[^A-Z0-9]+", "_", code_base.upper()).strip("_") or f"FASE_{i+1}"
        fases.append({
            "code": code,
            "label": short_label,
            "match_pattern": most_common,
        })

    # Extract WBS prefix for each processo (from first task's "no" field)
    processos_out = []
    for processo, tasks in uniform_processos.items():
        wbs_no = str(tasks[0].get("no", ""))
        # Prefix is everything except the last "." segment (e.g., "2.5.1" → "2.5")
        parts = wbs_no.split(".")
        wbs_prefix = ".".join(parts[:-1]) if len(parts) > 1 else wbs_no
        processos_out.append({
            "label": processo,
            "wbs_prefix": wbs_prefix,
        })

    # Sort processos by wbs_prefix (natural order in WBS)
    def wbs_key(p):
        parts = p["wbs_prefix"].split(".")
        return tuple(int(x) if x.isdigit() else x for x in parts)
    processos_out.sort(key=wbs_key)

    return {"processos": processos_out, "fases": fases}


def load_or_infer_matrix_structure(
    project_dir: Path,
    entries: list[dict],
    bar_titles: list[str],
    warnings: list,
    force_reinfer: bool = False,
) -> dict:
    """Loads matrix-structure.json if present, else infers and saves it.

    Returns a dict with at least {processos, fases, source}. When inference
    fails and no file exists, writes a "pending_user_input" stub with
    candidate_processos — the Claude Code agent is expected to complete it
    interactively via AskUserQuestion before the next run.
    """
    config_path = project_dir / "4-status-report" / "matrix-structure.json"

    if config_path.exists() and not force_reinfer:
        try:
            existing = json.loads(config_path.read_text(encoding="utf-8"))
            if existing.get("source") == "pending_user_input":
                warnings.append(
                    f"matrix-structure.json está aguardando input do usuário. "
                    f"Rode AskUserQuestion para confirmar processos e fases e atualize o arquivo."
                )
            return existing
        except json.JSONDecodeError as e:
            warnings.append(f"matrix-structure.json inválido ({e}), re-inferindo")

    inferred = infer_matrix_structure(entries, bar_titles)
    if inferred is None:
        # Emit stub for user to complete
        stub = {
            "version": 1,
            "source": "pending_user_input",
            "inferred_at": datetime.now().isoformat(timespec="seconds"),
            "candidate_processos": [{"label": t, "wbs_prefix": ""} for t in bar_titles],
            "processos": [],
            "fases": [],
            "_note": (
                "Inferência automática falhou. Preencha 'processos' (ordem das linhas) "
                "e 'fases' (ordem das colunas com match_pattern). Depois altere "
                "'source' para 'user_defined' e rode de novo."
            ),
        }
        config_path.parent.mkdir(parents=True, exist_ok=True)
        config_path.write_text(json.dumps(stub, indent=2, ensure_ascii=False), encoding="utf-8")
        warnings.append(
            f"matrix-structure.json criado como stub em {config_path}. "
            "Complete processos + fases via AskUserQuestion e rode de novo."
        )
        return stub

    structure = {
        "version": 1,
        "source": "inferred",
        "inferred_at": datetime.now().isoformat(timespec="seconds"),
        "processos": inferred["processos"],
        "fases": inferred["fases"],
    }
    config_path.parent.mkdir(parents=True, exist_ok=True)
    config_path.write_text(json.dumps(structure, indent=2, ensure_ascii=False), encoding="utf-8")
    return structure


def build_matrix_cells(
    structure: dict,
    entries: list[dict],
    report_date: datetime,
) -> dict:
    """Builds the processos × fases status grid using the provided structure.

    For each (processo, fase), finds the single task in `entries` whose `etapa`
    contains both the fase.match_pattern AND the processo.label (case-folded
    NFD-normalized). Uses the task's `status` field + fim_plan vs report_date
    to classify the cell: done, active (in_progress/blocked), overdue, not_started.

    Returns:
    {
        "processos": [...],  # from structure
        "fases": [...],      # from structure
        "matrix": [[cell, ...], ...],  # 2D grid [processo][fase]
        "meta": {"total_cells": int, "done_cells": int, "active_cells": int,
                 "overdue_cells": int, "not_started_cells": int}
    }
    """
    processos = structure.get("processos", [])
    fases = structure.get("fases", [])
    acoes = [e for e in entries if str(e.get("tipo", "")).lower() in ("ação", "acao")]

    matrix = []
    counters = {"done": 0, "active": 0, "overdue": 0, "not_started": 0, "missing": 0}
    for processo in processos:
        row = []
        p_norm = _normalize_match(processo.get("label", ""))
        for fase in fases:
            f_norm = _normalize_match(fase.get("match_pattern", ""))
            match = None
            for e in acoes:
                e_norm = _normalize_match(str(e.get("etapa", "")))
                if p_norm and f_norm and p_norm in e_norm and f_norm in e_norm:
                    match = e
                    break
            if not match:
                row.append({"status": "missing", "task_no": None, "deadline": None})
                counters["missing"] += 1
                continue
            status = match.get("status", "not_started")
            fim = match.get("fim_plan")
            deadline = fim.strftime("%d/%m") if isinstance(fim, datetime) else None
            if status == "done":
                cell_status = "done"
            elif status in ("in_progress", "blocked"):
                cell_status = "active"
            elif isinstance(fim, datetime) and fim < report_date and status != "done":
                cell_status = "overdue"
            else:
                cell_status = "not_started"
            row.append({
                "status": cell_status,
                "task_no": str(match.get("no", "")),
                "deadline": deadline,
            })
            counters[cell_status] += 1
        matrix.append(row)

    total = sum(v for k, v in counters.items() if k != "missing")
    return {
        "processos": processos,
        "fases": fases,
        "matrix": matrix,
        "meta": {
            "total_cells": total,
            "done_cells": counters["done"],
            "active_cells": counters["active"],
            "overdue_cells": counters["overdue"],
            "not_started_cells": counters["not_started"],
            "missing_cells": counters["missing"],
        },
    }


def compute_today_pct(milestones: list[dict], report_date: datetime) -> float | None:
    """Interpolates today's x-axis position (%) on the roadmap timeline using two
    milestone anchors (earliest and latest with known `date` + `left_pct`).

    Returns None if there are fewer than 2 usable anchor milestones.
    Clamps to [0.0, 100.0].
    """
    anchors = []
    for m in milestones:
        if m.get("date") and m.get("left_pct") is not None:
            try:
                dt = datetime.strptime(m["date"], "%Y-%m-%d")
                anchors.append((dt, float(m["left_pct"])))
            except (ValueError, TypeError):
                continue
    if len(anchors) < 2:
        return None
    anchors.sort(key=lambda a: a[0])
    first_dt, first_pct = anchors[0]
    last_dt, last_pct = anchors[-1]
    total_days = (last_dt - first_dt).days
    if total_days <= 0:
        return None
    delta = (report_date - first_dt).days
    pct = first_pct + (delta / total_days) * (last_pct - first_pct)
    return max(0.0, min(100.0, pct))


def parse_briefing(path: Path, warnings: list) -> dict:
    if not path.exists():
        warnings.append("BRIEFING.md não encontrado — project.goal pode estar incompleto")
        return {}
    text = path.read_text(encoding="utf-8")
    result = {}
    # Extract Objetivo section (first paragraph)
    m = re.search(r"^##\s+Objetivo\s*\n\s*\n?(.+?)(?=\n##|\Z)", text, re.DOTALL | re.MULTILINE | re.IGNORECASE)
    if m:
        para = m.group(1).strip().split("\n\n")[0]
        result["goal"] = para.strip()
    return result


def parse_claude_md(path: Path, warnings: list) -> dict:
    if not path.exists():
        return {}
    text = path.read_text(encoding="utf-8")
    result = {}
    m = re.search(r"clickup_list_id:\s*(\S+)", text)
    if m:
        result["clickup_list_id"] = m.group(1).strip()
    m = re.search(r"email:\s*(\S+@\S+)", text)
    if m:
        result["pm_email"] = m.group(1).strip()
    return result


# ---------- Synthesis (deterministic heuristics) ----------

# =============================================================================
# Status classification framework (v1.7) — PMI-flavored
# =============================================================================
# Replaces the legacy 3-zone heuristic that conflated risk-mapping with
# project-health. Now: 5 deterministic metrics from Cronograma.xlsx + 4 zones
# (🔵 Entregas Avançadas / 🟢 No Prazo / 🟡 Atenção / 🔴 Crítico) + gates.
# =============================================================================

def _filter_devido_hoje(actions: list[dict], report_date: datetime) -> list[dict]:
    """Tasks whose planned end is ≤ report_date (should have been delivered by now)."""
    return [
        a for a in actions
        if isinstance(a.get("fim_plan"), datetime) and a["fim_plan"] <= report_date
    ]


def _filter_iniciavel_hoje(actions: list[dict], report_date: datetime) -> list[dict]:
    """Tasks whose planned start is ≤ report_date (should be started by now)."""
    return [
        a for a in actions
        if isinstance(a.get("inicio_plan"), datetime) and a["inicio_plan"] <= report_date
    ]


def _filter_scope_ativo(actions: list[dict], report_date: datetime) -> list[dict]:
    """Tasks already in the radar (planned to start within 14 days)."""
    cutoff = report_date + timedelta(days=14)
    return [
        a for a in actions
        if isinstance(a.get("inicio_plan"), datetime) and a["inicio_plan"] <= cutoff
    ]


def _task_duration_days(a: dict) -> float:
    """Nominal duration of a task in calendar days. Returns 0 if endpoints missing.
    Min 1 day to avoid division-by-zero for same-day tasks."""
    ip, fp = a.get("inicio_plan"), a.get("fim_plan")
    if not (isinstance(ip, datetime) and isinstance(fp, datetime)):
        return 0.0
    return max(1.0, (fp - ip).days)


def metric_delivery_gap(actions: list[dict], report_date: datetime) -> float:
    """DG = (devido_hoje − done_até_hoje) / devido_hoje.

    Fraction of tasks that were due and not yet completed. Returns 0 when no
    task was due (early project) — a healthy signal, not a null.
    """
    devido = _filter_devido_hoje(actions, report_date)
    if not devido:
        return 0.0
    done = sum(1 for a in devido if a.get("status") == "done")
    return (len(devido) - done) / len(devido)


def metric_start_gap(actions: list[dict], report_date: datetime) -> float:
    """SG = late_start_count / iniciavel_hoje.

    Fraction of tasks that should have started but have no actual start date.
    """
    iniciaveis = _filter_iniciavel_hoje(actions, report_date)
    if not iniciaveis:
        return 0.0
    late = sum(
        1 for a in iniciaveis
        if not isinstance(a.get("inicio_real"), datetime) and a.get("status") == "not_started"
    )
    return late / len(iniciaveis)


def metric_spi(actions: list[dict], report_date: datetime) -> float:
    """SPI = EV / PV (PMI/PMBOK Schedule Performance Index).

    EV (Earned Value) = Σ duration of tasks already done
    PV (Planned Value) = Σ planned duration that should have been completed by now
      - For done tasks fully in past: full duration counts
      - For in-progress tasks whose end date has passed: full duration expected
      - For in-progress tasks whose end date is in the future: proportional
        (elapsed fraction of planned duration)

    SPI = 1.0 means on schedule. > 1.0 means ahead. < 1.0 means behind.
    """
    ev = 0.0
    pv = 0.0
    for a in actions:
        ip, fp = a.get("inicio_plan"), a.get("fim_plan")
        if not (isinstance(ip, datetime) and isinstance(fp, datetime)):
            continue
        dur = _task_duration_days(a)
        # EV: contribution if done
        if a.get("status") == "done":
            ev += dur
        # PV: how much of this task was supposed to be done by report_date
        if report_date >= fp:
            pv += dur  # full duration was supposed to be earned
        elif report_date <= ip:
            pv += 0.0  # task shouldn't have started yet
        else:
            elapsed = (report_date - ip).days
            pv += dur * (elapsed / dur) if dur > 0 else 0.0
    if pv == 0.0:
        # No work was supposed to be done yet — healthy neutral state
        return 1.0
    return ev / pv


def metric_msi(milestones: list[dict], report_date: datetime) -> int:
    """MSI = max slip days among MAJOR (gate) milestones that are overdue.

    Only counts milestones with `major=True`. Returns 0 when no major is
    overdue. A single 10-day slip on a critical gate is more signal than
    a diluted SPI, hence a separate metric.
    """
    max_slip = 0
    for m in milestones:
        if not m.get("major"):
            continue
        date_str = m.get("date")
        if not date_str:
            continue
        try:
            dt = datetime.strptime(date_str, "%Y-%m-%d")
        except ValueError:
            continue
        if m.get("status") == "overdue" or (dt < report_date and m.get("status") != "done"):
            slip = (report_date - dt).days
            if slip > max_slip:
                max_slip = slip
    return max_slip


def metric_edr(actions: list[dict]) -> float:
    """EDR = done_before_fim_plan / total_done.

    Among completed tasks, fraction delivered strictly before the planned end date.
    """
    done = [a for a in actions if a.get("status") == "done"]
    if not done:
        return 0.0
    early = 0
    for a in done:
        fr = a.get("fim_real")
        fp = a.get("fim_plan")
        if isinstance(fr, datetime) and isinstance(fp, datetime) and fr < fp:
            early += 1
    return early / len(done)


def metric_ahead_ratio(actions: list[dict], report_date: datetime) -> float:
    """Ahead ratio = done_early / total_due_so_far.

    Distinct from EDR because it uses TOTAL_DUE as denominator (not total done).
    Gate threshold for celebratory 🔵 override: ≥ 25% means project is genuinely
    running ahead, not just lucky with a small sample.
    """
    due = _filter_devido_hoje(actions, report_date)
    if not due:
        return 0.0
    early = 0
    for a in due:
        fr = a.get("fim_real")
        fp = a.get("fim_plan")
        if isinstance(fr, datetime) and isinstance(fp, datetime) and fr < fp:
            early += 1
    return early / len(due)


# ----- Zone classification -----

_ZONE_ORDER = {"blue": 0, "green": 1, "yellow": 2, "red": 3}


def _worst(zones: list[str]) -> str:
    """Returns the worst (highest rank) zone among inputs."""
    if not zones:
        return "green"
    return max(zones, key=lambda z: _ZONE_ORDER.get(z, 0))


def _zone_of_dg(dg: float) -> str:
    if dg <= 0.001: return "blue"
    if dg <= 0.05:  return "green"
    if dg <= 0.15:  return "yellow"
    return "red"


def _zone_of_sg(sg: float) -> str:
    if sg <= 0.001: return "blue"
    if sg <= 0.10:  return "green"
    if sg <= 0.25:  return "yellow"
    return "red"


def _zone_of_spi(spi: float) -> str:
    if spi >= 1.05: return "blue"
    if spi >= 0.95: return "green"
    if spi >= 0.85: return "yellow"
    return "red"


def _zone_of_msi(msi_days: int) -> str:
    if msi_days == 0: return "blue"
    if msi_days <= 7: return "yellow"
    return "red"


def classify_zone(metrics: dict, milestones: list[dict], report_date: datetime) -> dict:
    """Applies the worst-of rule across metric zones + gate overrides.

    Returns:
        {
            "zone": "blue|green|yellow|red",
            "metric_zones": {"dg": "blue", "sg": "blue", "spi": "green", ...},
            "reasons": ["DG = 0.0% (🔵)", "SG = 0.0% (🔵)", ...],
            "gates_fired": ["ahead_ratio override (≥25%)"],
        }
    """
    dg, sg, spi = metrics["dg"], metrics["sg"], metrics["spi"]
    msi, edr = metrics["msi"], metrics["edr"]
    ahead = metrics.get("ahead_ratio", 0.0)

    metric_zones = {
        "dg": _zone_of_dg(dg),
        "sg": _zone_of_sg(sg),
        "spi": _zone_of_spi(spi),
        "msi": _zone_of_msi(msi),
    }
    base_zone = _worst(list(metric_zones.values()))

    reasons = [
        f"DG = {dg*100:.1f}% ({_emoji(metric_zones['dg'])})",
        f"SG = {sg*100:.1f}% ({_emoji(metric_zones['sg'])})",
        f"SPI = {spi:.2f} ({_emoji(metric_zones['spi'])})",
        f"MSI = {msi}d ({_emoji(metric_zones['msi'])})",
        f"EDR = {edr*100:.0f}%",
    ]
    gates_fired = []

    # Gate RED: terminal milestone overdue
    if milestones:
        major_dated = [
            m for m in milestones
            if m.get("major") and m.get("date")
        ]
        if major_dated:
            # Terminal = last major by date
            def _key(m):
                try:
                    return datetime.strptime(m["date"], "%Y-%m-%d")
                except (ValueError, TypeError):
                    return datetime.min
            terminal = max(major_dated, key=_key)
            terminal_dt = _key(terminal)
            if terminal_dt < report_date and terminal.get("status") != "done":
                gates_fired.append(f"Gate RED: marco terminal {terminal.get('label')} atrasado")
                reasons.append(f"🔴 Gate: marco terminal {terminal.get('label')} em atraso")
                return {
                    "zone": "red",
                    "metric_zones": metric_zones,
                    "reasons": reasons,
                    "gates_fired": gates_fired,
                }

    # Gate BLUE (celebratory): ahead_ratio ≥ 25%
    if ahead >= 0.25 and base_zone in ("blue", "green"):
        gates_fired.append(f"Gate BLUE: ahead_ratio = {ahead*100:.0f}% ≥ 25%")
        reasons.append(f"🔵 Override: ahead_ratio {ahead*100:.0f}% (≥ 25%)")
        return {
            "zone": "blue",
            "metric_zones": metric_zones,
            "reasons": reasons,
            "gates_fired": gates_fired,
        }

    return {
        "zone": base_zone,
        "metric_zones": metric_zones,
        "reasons": reasons,
        "gates_fired": gates_fired,
    }


def _emoji(zone: str) -> str:
    return {"blue": "🔵", "green": "🟢", "yellow": "🟡", "red": "🔴"}.get(zone, "⚪")


# =============================================================================
# End of v1.7 status framework
# =============================================================================


def synthesize(
    entries: list[dict],
    risks: list[dict],
    changelog: list[dict],
    milestones: list[dict],
    report_date: datetime,
    warnings: list,
) -> dict:
    actions = [e for e in entries if str(e.get("tipo", "")).lower() in ("ação", "acao")]
    phases = [e for e in entries if str(e.get("tipo", "")).lower() == "fase"]

    total = len(actions)
    done = sum(1 for a in actions if a.get("status") == "done")
    overdue = sum(
        1 for a in actions
        if a.get("status") != "done"
        and a.get("fim_plan")
        and isinstance(a["fim_plan"], datetime)
        and a["fim_plan"] < report_date
    )
    pct_done = round(100 * done / total) if total else 0

    # Status overall — v1.7 PMI-flavored framework (5 metrics + 4 zones + gates).
    # Risks are NO LONGER considered here: a mapped risk is not a materialized
    # health issue. Risks show up in the dedicated OPR risks zone (v1.6) when
    # flagged as `is_incurred=True`.
    metrics = {
        "dg": metric_delivery_gap(actions, report_date),
        "sg": metric_start_gap(actions, report_date),
        "spi": metric_spi(actions, report_date),
        "msi": metric_msi(milestones, report_date),
        "edr": metric_edr(actions),
        "ahead_ratio": metric_ahead_ratio(actions, report_date),
    }
    classification = classify_zone(metrics, milestones, report_date)
    overall = classification["zone"]  # "blue" | "green" | "yellow" | "red"

    # Active sprint = first Fase not done with Fim Planejado >= report_date (ou a mais antiga em andamento)
    active_sprint = None
    active_idx = 0
    for i, phase in enumerate(sorted(phases, key=lambda p: p.get("inicio_plan") or datetime.max), 1):
        if phase.get("status") != "done":
            active_sprint = phase
            active_idx = i
            break

    # Highlights
    window_days = 14
    cutoff = report_date - timedelta(days=window_days)
    candidates = []
    for action in actions:
        if action.get("status") == "done" and isinstance(action.get("fim_real"), datetime) and action["fim_real"] >= cutoff:
            candidates.append({
                "text": f"Concluiu: {truncate(action.get('etapa', ''), 90)}",
                "date": action["fim_real"],
                "priority": 10,
            })
    for phase in phases:
        if isinstance(phase.get("fim_real"), datetime) and phase["fim_real"] >= cutoff:
            candidates.append({
                "text": f"Atingiu marco: {truncate(phase.get('etapa', ''), 80)}",
                "date": phase["fim_real"],
                "priority": 15,
            })
    keywords = ("decidido", "aprovado", "definido", "validado")
    for entry in changelog:
        if entry.get("op") == "comment" and any(k in entry.get("text", "").lower() for k in keywords):
            candidates.append({
                "text": f"Decidiu: {truncate(entry['text'], 90)}",
                "date": datetime.fromisoformat(entry["timestamp"]),
                "priority": 8,
            })
    candidates.sort(key=lambda c: (-c["priority"], -c["date"].timestamp()))
    highlights = [c["text"] for c in candidates[:5]]

    # Progress — structured list of concluded tasks in period (for OPR v1.6 layout)
    window_days_progress = 14
    cutoff_progress = report_date - timedelta(days=window_days_progress)
    progress_concluidas = []
    for action in actions:
        if action.get("status") == "done" and isinstance(action.get("fim_real"), datetime) and action["fim_real"] >= cutoff_progress:
            progress_concluidas.append({
                "wbs": str(action.get("no", "")),
                "text": truncate(action.get("etapa", ""), 60),
                "date": action["fim_real"].strftime("%d/%m"),
                "_date": action["fim_real"],
            })
    for phase in phases:
        if isinstance(phase.get("fim_real"), datetime) and phase["fim_real"] >= cutoff_progress:
            progress_concluidas.append({
                "wbs": str(phase.get("no", "")),
                "text": f"✓ {truncate(phase.get('etapa', ''), 55)}",
                "date": phase["fim_real"].strftime("%d/%m"),
                "_date": phase["fim_real"],
            })
    progress_concluidas.sort(key=lambda c: c.get("_date") or datetime.min, reverse=True)
    for item in progress_concluidas:
        item.pop("_date", None)
    progress_concluidas = progress_concluidas[:5]

    # Next steps
    lookahead = report_date + timedelta(days=14)
    next_candidates = []
    for phase in phases:
        ip = phase.get("inicio_plan")
        if isinstance(ip, datetime) and report_date <= ip <= lookahead:
            next_candidates.append({
                "action": f"Iniciar {truncate(phase.get('etapa', ''), 80)}",
                "deadline": ip.strftime("%d/%m"),
                "deadline_full": ip.strftime("%Y-%m-%d"),
                "rationale": f"Gate de entrada da fase {phase.get('no', '')}",
                "priority": 15,
            })
    for action in actions:
        status = action.get("status")
        fp = action.get("fim_plan")
        if status == "blocked":
            next_candidates.append({
                "action": f"Desbloquear: {truncate(action.get('etapa', ''), 80)}",
                "deadline": fp.strftime("%d/%m") if isinstance(fp, datetime) else "—",
                "deadline_full": fp.strftime("%Y-%m-%d") if isinstance(fp, datetime) else None,
                "rationale": "Bloqueio ativo — impede progresso",
                "priority": 20,
            })
        elif status != "done" and isinstance(fp, datetime) and report_date <= fp <= lookahead:
            next_candidates.append({
                "action": truncate(action.get("etapa", ""), 90),
                "deadline": fp.strftime("%d/%m"),
                "deadline_full": fp.strftime("%Y-%m-%d"),
                "rationale": f"Responsável: {action.get('responsavel', '—')}",
                "priority": 10,
            })
    next_candidates.sort(
        key=lambda c: (-c["priority"], c.get("deadline_full") or "9999")
    )
    next_steps = next_candidates[:5]

    # Progress — structured list of upcoming tasks (for OPR v1.6 layout).
    # Uses the same candidates as next_steps but with {wbs, text, date} schema.
    progress_proximas = []
    for phase in phases:
        ip = phase.get("inicio_plan")
        if isinstance(ip, datetime) and report_date <= ip <= lookahead:
            progress_proximas.append({
                "wbs": str(phase.get("no", "")),
                "text": f"Iniciar {truncate(phase.get('etapa', ''), 55)}",
                "date": ip.strftime("%d/%m"),
                "_date": ip,
                "_priority": 15,
            })
    for action in actions:
        status = action.get("status")
        fp = action.get("fim_plan")
        if status == "blocked":
            progress_proximas.append({
                "wbs": str(action.get("no", "")),
                "text": f"⚠ Desbloquear: {truncate(action.get('etapa', ''), 50)}",
                "date": fp.strftime("%d/%m") if isinstance(fp, datetime) else "—",
                "_date": fp if isinstance(fp, datetime) else datetime.max,
                "_priority": 20,
            })
        elif status != "done" and isinstance(fp, datetime) and report_date <= fp <= lookahead:
            progress_proximas.append({
                "wbs": str(action.get("no", "")),
                "text": truncate(action.get("etapa", ""), 60),
                "date": fp.strftime("%d/%m"),
                "_date": fp,
                "_priority": 10,
            })
    progress_proximas.sort(key=lambda c: (-c.get("_priority", 0), c.get("_date") or datetime.max))
    for item in progress_proximas:
        item.pop("_date", None)
        item.pop("_priority", None)
    progress_proximas = progress_proximas[:5]

    # Attentions — prefer authoritative severity_class from riscos.html;
    # skip upsides (codes starting with O = Oportunidade, not risk to watch).
    attentions = []
    for risk in risks:
        if risk.get("is_upside"):
            continue
        sc = risk.get("severity_class") or ""
        if sc == "crit":
            sev = "critical"
        elif sc == "high":
            sev = "warning"
        else:
            # Fallback: derive from prob×impact for legacy data without severity_class
            sev = _risk_severity(risk.get("probability", ""), risk.get("impact", ""))
        if sev in ("critical", "warning"):
            attentions.append({
                "severity": sev,
                "text": f"{risk['code']} — {truncate(risk['title'], 90)}",
                "source": "risk",
            })
    for action in actions:
        if action.get("status") == "blocked":
            attentions.append({
                "severity": "critical",
                "text": f"Bloqueado: {truncate(action.get('etapa', ''), 85)}",
                "source": "blocked",
            })
    if total and overdue:
        pct_overdue = round(100 * overdue / total)
        attentions.append({
            "severity": "warning" if pct_overdue <= 20 else "critical",
            "text": f"{overdue} ação(ões) atrasada(s) ({pct_overdue}% do total)",
            "source": "overdue_summary",
        })
    order = {"critical": 0, "warning": 1, "neutral": 2}
    attentions.sort(key=lambda a: order.get(a["severity"], 99))
    attentions = attentions[:5]

    # Macro milestones — prefer gates M0-M7 parsed from roadmap-marcos.html (canonical
    # source of visual truth from building-project-plan). Falls back to xlsx root phases
    # only if the HTML artifact is absent or doesn't have a .milestones lane.
    root_phases = [p for p in phases if p.get("no") and "." not in str(p["no"])]
    root_phases.sort(key=lambda p: p.get("inicio_plan") or datetime.max)

    if milestones:
        # Show all milestones up to 8 (full M0-M7 strip like the Paper artboard).
        # If there are more than 8, filter to .gate (major) to avoid clutter.
        chosen = milestones if len(milestones) <= 8 else [m for m in milestones if m.get("major")]
        macro_milestones = [
            {
                "label": truncate(m.get("label", ""), 22),
                "status": m.get("status", "not_started"),
                "date": m.get("date"),
                "date_short": m.get("date_short"),
                "description": m.get("description", ""),
                "code": m.get("code", ""),
                "major": m.get("major", False),
                "left_pct": m.get("left_pct"),
            }
            for m in chosen[:8]
        ]
    else:
        macro_milestones = []
        for p in root_phases[:7]:
            status = "not_started"
            if p.get("status") == "done":
                status = "done"
            elif isinstance(p.get("fim_plan"), datetime) and p["fim_plan"] < report_date and p.get("status") != "done":
                status = "overdue"
            elif p.get("status") == "in_progress":
                status = "in_progress"
            macro_milestones.append({
                "label": truncate(p.get("etapa", ""), 28),
                "status": status,
                "no": p.get("no"),
            })

    # Sprints — simplified: treat root phases as sprints
    sprints = []
    for i, p in enumerate(root_phases, 1):
        period = ""
        if isinstance(p.get("inicio_plan"), datetime) and isinstance(p.get("fim_plan"), datetime):
            period = f"{p['inicio_plan'].strftime('%d/%b')} → {p['fim_plan'].strftime('%d/%b')}"
        sprint_status = "future"
        if active_sprint and p.get("no") == active_sprint.get("no"):
            sprint_status = "active"
        elif p.get("status") == "done":
            sprint_status = "done"
        sprints.append({
            "code": f"S{i - 1}",
            "title": p.get("etapa", ""),
            "period_label": period,
            "status": sprint_status,
            "no": p.get("no"),
            "fronts": [],
        })

    # Hero sentences
    hero_executive = f"{done} de {total} tarefas concluídas ({pct_done}%)" if total else "Sem ações cadastradas"
    active_count = sum(1 for s in sprints if s["status"] == "active")
    sprint_progress = f"{active_count} de {len(sprints)} sprint(s) em execução" if sprints else "Sem sprints"
    # Exclude upsides (O01, O02...) from risk counting — they are opportunities, not threats
    actual_risks = [r for r in risks if not r.get("is_upside")]
    critical_risks = sum(1 for r in actual_risks if r.get("severity_class") == "crit")
    high_risks = sum(1 for r in actual_risks if r.get("severity_class") == "high")
    if not actual_risks:
        risks_sentence = "Nenhum risco mapeado"
    elif critical_risks:
        risks_sentence = f"{len(actual_risks)} risco(s) mapeado(s) — {critical_risks} crítico(s)"
    elif high_risks:
        risks_sentence = f"{len(actual_risks)} risco(s) mapeado(s) — {high_risks} alto(s)"
    else:
        risks_sentence = f"{len(actual_risks)} risco(s) mapeado(s) — severidade média/baixa"

    return {
        "overall": overall,
        "metrics": metrics,
        "metric_zones": classification["metric_zones"],
        "status_reasons": classification["reasons"],
        "gates_fired": classification["gates_fired"],
        "percent_done": pct_done,
        "total_actions": total,
        "done_actions": done,
        "overdue_actions": overdue,
        "active_sprint": (
            {
                "code": f"S{next((i for i, p in enumerate(root_phases) if p.get('no') == active_sprint.get('no')), 0)}",
                "no": active_sprint.get("no"),
                "title": active_sprint.get("etapa"),
                "phase_name": (active_sprint.get("etapa") or "").upper()[:24],
                "eyebrow": f"S{active_idx - 1} · {(active_sprint.get('etapa') or '').upper()[:22]}",
                "period_label": _format_period(active_sprint),
            }
            if active_sprint else None
        ),
        "active_sprint_index": active_idx,
        "total_sprints": len(sprints),
        "hero_sentence": hero_executive,
        "sprint_progress_sentence": sprint_progress,
        "risks_sentence": risks_sentence,
        "_highlights": highlights,
        "_next_steps": next_steps,
        "_progress_concluidas": progress_concluidas,
        "_progress_proximas": progress_proximas,
        "_attentions": attentions,
        "_macro_milestones": macro_milestones,
        "_sprints": sprints,
    }


def _format_period(phase: dict) -> str:
    ip = phase.get("inicio_plan")
    fp = phase.get("fim_plan")
    if isinstance(ip, datetime) and isinstance(fp, datetime):
        return f"{ip.strftime('%d/%b')} → {fp.strftime('%d/%b')}"
    return ""


def _risk_severity(prob: str, impact: str) -> str:
    if prob == "alta" and impact in ("critico", "alto"):
        return "critical"
    if prob in ("alta", "media") or impact in ("critico", "alto"):
        return "warning"
    return "neutral"


def truncate(s: str, max_len: int) -> str:
    s = (s or "").strip()
    return s if len(s) <= max_len else s[: max_len - 1].rstrip() + "…"


# ---------- Main ----------

def collect(project_dir: Path, report_date_str: str) -> dict:
    report_date = datetime.strptime(report_date_str, "%Y-%m-%d")
    warnings: list = []

    xlsx = project_dir / "4-status-report" / "Cronograma.xlsx"
    if not xlsx.exists():
        raise FileNotFoundError(
            f"Cronograma LIVE não inicializado: {xlsx}\n"
            f"Rode `managing-action-plan init` primeiro."
        )

    entries = read_cronograma_live(xlsx, default_year=report_date.year, warnings=warnings)
    changelog = read_changelog(project_dir / "4-status-report" / "changelog.md")
    sync_state = read_sync_state(
        project_dir / "4-status-report" / ".sync-state.json",
        report_date,
        warnings,
    )
    plano_data = parse_plano_projeto(project_dir / "1-planning" / "plano-projeto.html", warnings)
    briefing = parse_briefing(project_dir / "BRIEFING.md", warnings)
    claude_data = parse_claude_md(project_dir / "CLAUDE.md", warnings)
    risks = parse_risks(project_dir / "1-planning" / "artefatos" / "riscos.html", warnings)
    roadmap_html_path = project_dir / "1-planning" / "artefatos" / "roadmap-marcos.html"
    milestones = parse_milestones(
        roadmap_html_path,
        warnings,
        default_year=report_date.year,
        report_date=report_date,
    )

    # Roadmap overlays — inputs for build_pptx to inject into the HTML screenshot:
    #   - today_pct: where to draw the vertical "HOJE" line (interpolated from M0↔M7 anchors)
    #   - bar_status_by_title: per-bar status derived from matching title ↔ Cronograma.xlsx etapa
    bar_titles = _collect_bar_titles(roadmap_html_path)
    today_pct = compute_today_pct(milestones, report_date)
    bar_status_map = {
        title: aggregate_bar_status(title, entries, report_date)
        for title in bar_titles
    }
    matched_count = sum(1 for s in bar_status_map.values() if s != "future")
    if bar_titles and matched_count == 0:
        warnings.append(
            f"Coloração de bars: 0 de {len(bar_titles)} títulos casaram com etapas do Cronograma.xlsx. "
            "Bars ficarão todas cinza. Confira consistência entre títulos do roadmap e nomes das ações."
        )
    roadmap_overlays = {
        "today_pct": today_pct,
        "today_label": f"HOJE · {report_date.strftime('%d/%m')}",
        "bar_status_by_title": bar_status_map,
    }

    # Processos × Fases matrix for the "Visão Geral do Roadmap" slide.
    # Reads or infers structure; persists to <proj>/4-status-report/matrix-structure.json
    matrix_structure_cfg = load_or_infer_matrix_structure(
        project_dir, entries, bar_titles, warnings,
    )
    if matrix_structure_cfg.get("source") == "pending_user_input":
        # Stub exists → slide 3 will render empty; other slides OK
        roadmap_structure = {
            "processos": [],
            "fases": [],
            "matrix": [],
            "meta": {"total_cells": 0, "done_cells": 0, "active_cells": 0,
                     "overdue_cells": 0, "not_started_cells": 0, "missing_cells": 0},
            "source": "pending_user_input",
            "candidate_processos": matrix_structure_cfg.get("candidate_processos", []),
        }
    else:
        roadmap_structure = build_matrix_cells(matrix_structure_cfg, entries, report_date)
        roadmap_structure["source"] = matrix_structure_cfg.get("source", "inferred")

    synth = synthesize(entries, risks, changelog, milestones, report_date, warnings)

    clickup_url = None
    list_id = sync_state.get("clickup_list_id") or claude_data.get("clickup_list_id")
    if list_id:
        clickup_url = f"https://app.clickup.com/{list_id}/l/{list_id}"

    result = CollectResult(
        report_date=report_date_str,
        project={
            "name": plano_data.get("name") or project_dir.name,
            "goal": briefing.get("goal") or plano_data.get("goal") or "",
            "pm": plano_data.get("pm") or "",
            "pm_email": claude_data.get("pm_email") or "",
            "period_label": _period_label(report_date),
            "footer_label": f"M7 Investimentos · {MONTH_PT[report_date.month]} {report_date.year}",
            "clickup_list_url": clickup_url,
        },
        status={
            k: v for k, v in synth.items()
            if not k.startswith("_")
        },
        highlights=synth["_highlights"],
        next_steps=synth["_next_steps"],
        progress_concluidas=synth["_progress_concluidas"],
        progress_proximas=synth["_progress_proximas"],
        attentions=synth["_attentions"],
        milestones=milestones,
        macro_milestones=synth["_macro_milestones"],
        sprints=synth["_sprints"],
        fronts=[],
        sprint_actions=[
            {
                "no": a.get("no"),
                "etapa": truncate(a.get("etapa", ""), 80),
                "status": a.get("status"),
                "responsavel": a.get("responsavel"),
                "fim_planejado": a["fim_plan"].strftime("%Y-%m-%d") if isinstance(a.get("fim_plan"), datetime) else None,
                "clickup_id": a.get("clickup_id"),
            }
            for a in [e for e in entries if str(e.get("tipo", "")).lower() in ("ação", "acao")
                      and e.get("status") in ("in_progress", "blocked")][:10]
        ],
        risks=risks,
        changelog_entries=[{k: v for k, v in e.items() if k != "raw"} for e in changelog[:30]],
        roadmap_overlays=roadmap_overlays,
        roadmap_structure=roadmap_structure,
        warnings=warnings,
    )
    return result.to_dict()


def _period_label(report_date: datetime, span_days: int = 14) -> str:
    start = report_date - timedelta(days=span_days)
    return f"Quinzena {start.strftime('%d/%m')} — {report_date.strftime('%d/%m/%Y')}"


def main():
    p = argparse.ArgumentParser(description="Collect status report data deterministically")
    p.add_argument("--project-dir", required=True, type=Path)
    p.add_argument("--report-date", required=True, help="YYYY-MM-DD")
    p.add_argument("--output", type=Path, help="Escrever JSON aqui; stdout se omitido")
    args = p.parse_args()

    if not args.project_dir.exists():
        print(f"✗ project-dir não encontrado: {args.project_dir}", file=sys.stderr)
        sys.exit(2)
    if not (args.project_dir / "CLAUDE.md").exists():
        print(
            "✗ Projeto não inicializado: CLAUDE.md ausente.\n"
            "  Rode `initializing-project` primeiro.",
            file=sys.stderr,
        )
        sys.exit(2)

    try:
        data = collect(args.project_dir, args.report_date)
    except FileNotFoundError as e:
        print(f"✗ {e}", file=sys.stderr)
        sys.exit(2)
    except ValueError as e:
        print(f"✗ {e}", file=sys.stderr)
        sys.exit(2)

    payload = json.dumps(data, indent=2, default=str, ensure_ascii=False)
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(payload, encoding="utf-8")
        print(f"✓ Coleta gravada em {args.output}")
        if data["warnings"]:
            print(f"⚠ {len(data['warnings'])} warning(s):", file=sys.stderr)
            for w in data["warnings"]:
                print(f"   · {w}", file=sys.stderr)
    else:
        print(payload)


if __name__ == "__main__":
    main()
